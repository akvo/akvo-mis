import re
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl

from api.v1.v1_forms.constants import FormTypes
from api.v1.v1_forms.services.xlsform_export import (
    _DEFAULT_LANG_CODE,
    _extract_iso,
)

# Supported XLSForm types for Akvo MIS
_SUPPORTED_BASE_TYPES = {
    "text",
    "integer",
    "decimal",
    "date",
    "geopoint",
    "image",
    "file",
    "begin_group",
    "begin group",
    "begin_repeat",
    "begin repeat",
    "end_group",
    "end group",
    "end_repeat",
    "end repeat",
}


def parse_relevant_expression(
    expr: str, name_to_tmp_id: Dict[str, int]
) -> Tuple[List[Dict[str, Any]], str, Optional[str]]:
    """Parse an XLSForm 'relevant' XPath expression back into dependency dicts.

    Reverses `_build_relevant_expression()` from `xlsform_export.py`.
    Supported conditions:
      - selected(${name}, 'val')
      - selected(${name}, 'a') or selected(${name}, 'b')
      - ${name} >= N
      - ${name} <= N
      - ${name} = 'val'
      - ${name} != 'val' and string-length(${name}) > 0

    Returns:
      (dependencies, dependency_rule, warning_message_or_None)
    """
    if not expr or not isinstance(expr, str) or not expr.strip():
        return [], "AND", None

    clean_expr = expr.strip()

    # Determine overall rule: OR vs AND
    # Look for top-level ' or ' outside parentheses
    has_top_level_or = False
    depth = 0
    i = 0
    while i < len(clean_expr):
        ch = clean_expr[i]
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        elif depth == 0:
            if clean_expr[i : i + 4].lower() == " or ":  # noqa
                has_top_level_or = True
                break
        i += 1

    rule = "OR" if has_top_level_or else "AND"

    # Regex patterns for matching atomic sub-clauses
    pattern_selected = re.compile(
        r"selected\(\s*\$\{([^}]+)\}\s*,\s*['\"]([^'\"]+)['\"]\s*\)",
        re.IGNORECASE,
    )
    pattern_gte = re.compile(r"\$\{([^}]+)\}\s*>=\s*(-?[0-9]+(?:\.[0-9]+)?)")
    pattern_lte = re.compile(r"\$\{([^}]+)\}\s*<=\s*(-?[0-9]+(?:\.[0-9]+)?)")
    pattern_eq = re.compile(r"\$\{([^}]+)\}\s*=\s*['\"]([^'\"]*)['\"]")
    pattern_neq = re.compile(
        r"\$\{([^}]+)\}\s*!=\s*['\"]([^'\"]*)['\"]"
        r"(?:\s+and\s+string-length\(\s*\$\{\1\}\s*\)\s*>\s*0)?"
    )

    clauses: List[str] = []
    if rule == "OR":
        depth = 0
        last_idx = 0
        for idx in range(len(clean_expr)):
            if clean_expr[idx] == "(":
                depth += 1
            elif clean_expr[idx] == ")":
                depth = max(0, depth - 1)
            elif (
                depth == 0
                and clean_expr[idx : idx + 4].lower() == " or "  # noqa
            ):
                clauses.append(clean_expr[last_idx:idx].strip())
                last_idx = idx + 4
        clauses.append(clean_expr[last_idx:].strip())
    else:
        depth = 0
        last_idx = 0
        for idx in range(len(clean_expr)):
            if clean_expr[idx] == "(":
                depth += 1
            elif clean_expr[idx] == ")":
                depth = max(0, depth - 1)
            elif (
                depth == 0
                and clean_expr[idx : idx + 5].lower() == " and "  # noqa
            ):
                clauses.append(clean_expr[last_idx:idx].strip())
                last_idx = idx + 5
        clauses.append(clean_expr[last_idx:].strip())

    dependencies: List[Dict[str, Any]] = []

    for clause in clauses:
        clause_clean = clause.strip()
        while clause_clean.startswith("(") and clause_clean.endswith(")"):
            depth = 0
            enclosing = True
            for ci, c in enumerate(clause_clean[:-1]):
                if c == "(":
                    depth += 1
                elif c == ")":
                    depth -= 1
                if depth == 0 and ci > 0:
                    enclosing = False
                    break
            if enclosing:
                clause_clean = clause_clean[1:-1].strip()
            else:
                break

        selected_matches = pattern_selected.findall(clause_clean)
        if selected_matches:
            test_sub = pattern_selected.sub("", clause_clean)
            test_sub_clean = re.sub(
                r"\s*or\s*", "", test_sub, flags=re.IGNORECASE
            ).strip()
            if not test_sub_clean:
                by_var: Dict[str, List[str]] = {}
                for v_name, opt_val in selected_matches:
                    by_var.setdefault(v_name, []).append(opt_val)

                for v_name, opt_vals in by_var.items():
                    if v_name not in name_to_tmp_id:
                        return (
                            [],
                            rule,
                            f"Target question '{v_name}' not found",
                        )
                    q_id = name_to_tmp_id[v_name]
                    dependencies.append({"id": q_id, "options": opt_vals})
                continue

        m_gte = pattern_gte.fullmatch(clause_clean)
        if m_gte:
            v_name, val_str = m_gte.groups()
            if v_name not in name_to_tmp_id:
                return (
                    [],
                    rule,
                    f"Target question '{v_name}' not found",
                )
            num_val = float(val_str) if "." in val_str else int(val_str)
            q_id = name_to_tmp_id[v_name]
            found = False
            for dep in dependencies:
                if (
                    dep["id"] == q_id
                    and "min" not in dep
                    and "options" not in dep
                ):
                    dep["min"] = num_val
                    found = True
                    break
            if not found:
                dependencies.append({"id": q_id, "min": num_val})
            continue

        m_lte = pattern_lte.fullmatch(clause_clean)
        if m_lte:
            v_name, val_str = m_lte.groups()
            if v_name not in name_to_tmp_id:
                return (
                    [],
                    rule,
                    f"Target question '{v_name}' not found",
                )
            num_val = float(val_str) if "." in val_str else int(val_str)
            q_id = name_to_tmp_id[v_name]
            found = False
            for dep in dependencies:
                if (
                    dep["id"] == q_id
                    and "max" not in dep
                    and "options" not in dep
                ):
                    dep["max"] = num_val
                    found = True
                    break
            if not found:
                dependencies.append({"id": q_id, "max": num_val})
            continue

        m_eq = pattern_eq.fullmatch(clause_clean)
        if m_eq:
            v_name, eq_val = m_eq.groups()
            if v_name not in name_to_tmp_id:
                return (
                    [],
                    rule,
                    f"Target question '{v_name}' not found",
                )
            q_id = name_to_tmp_id[v_name]
            dependencies.append({"id": q_id, "equal": eq_val})
            continue

        m_neq = pattern_neq.fullmatch(clause_clean)
        if m_neq:
            v_name, neq_val = m_neq.groups()
            if v_name not in name_to_tmp_id:
                return (
                    [],
                    rule,
                    f"Target question '{v_name}' not found",
                )
            q_id = name_to_tmp_id[v_name]
            dependencies.append({"id": q_id, "notEqual": neq_val})
            continue

        pattern_str_len = re.compile(
            r"string-length\(\s*\$\{([^}]+)\}\s*\)\s*>\s*0"
        )
        m_strlen = pattern_str_len.fullmatch(clause_clean)
        if m_strlen:
            v_name = m_strlen.group(1)
            if v_name in name_to_tmp_id:
                continue

        return (
            [],
            rule,
            f"Unrecognized relevant expression clause '{clause_clean}'",
        )

    if not dependencies:
        return [], rule, f"Could not extract dependencies from '{expr}'"

    return dependencies, rule, None


def _strip_outer_parens(s: str) -> str:
    """Recursively strips matching outermost parentheses."""
    s = s.strip()
    while s.startswith("(") and s.endswith(")"):
        depth = 0
        matched = False
        for i, ch in enumerate(s):
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    if i == len(s) - 1:
                        matched = True
                    break
        if matched:
            s = s[1:-1].strip()
        else:
            break
    return s


def _parse_constraint(
    constraint_str: str, q_name: Optional[str] = None
) -> Tuple[Dict[str, Any], Optional[str]]:
    """Parse XLSForm constraint expression into rule min/max.

    Supports:
      - . <= 100, (. <= 100), ((. <= 100))
      - ${q_name} <= 100, (${q_name} <= 100)
      - . >= 0 and . <= 100
      - (. >= 0 and . <= 100), ((. >= 0) and (. <= 100))
      - . <= 100 and . >= 0
      - 0 <= . and . <= 100
      - < and > boundaries
      - Floats and integers

    Returns:
      (rule_dict, warning_message_or_None)
    """
    if not constraint_str or not isinstance(constraint_str, str):
        return {}, None

    clean = constraint_str.strip()
    if not clean:
        return {}, None

    clean = _strip_outer_parens(clean)

    # Split by top-level ' and ' or ' AND '
    clauses = []
    depth = 0
    last_idx = 0
    for idx in range(len(clean)):
        if clean[idx] == "(":
            depth += 1
        elif clean[idx] == ")":
            depth = max(0, depth - 1)
        elif depth == 0 and clean[idx:idx + 5].lower() == " and ":
            clauses.append(clean[last_idx:idx].strip())
            last_idx = idx + 5
    clauses.append(clean[last_idx:].strip())

    rule: Dict[str, Any] = {}
    unparsed_clauses = []

    var_pat = r"(?:\.|\$\{[^}]+\})"
    num_pat = r"(-?[0-9]+(?:\.[0-9]+)?)"

    for raw_clause in clauses:
        clause = _strip_outer_parens(raw_clause)
        if not clause:
            continue

        matched = False

        # Pattern 1: target >= N or target > N
        m1 = re.match(rf"^{var_pat}\s*(>=|>)\s*{num_pat}$", clause)
        if m1:
            op, val_str = m1.groups()
            val = float(val_str) if "." in val_str else int(val_str)
            rule["min"] = val
            matched = True

        # Pattern 2: target <= N or target < N
        if not matched:
            m2 = re.match(rf"^{var_pat}\s*(<=|<)\s*{num_pat}$", clause)
            if m2:
                op, val_str = m2.groups()
                val = float(val_str) if "." in val_str else int(val_str)
                rule["max"] = val
                matched = True

        # Pattern 3: N <= target or N < target (means target >= N)
        if not matched:
            m3 = re.match(rf"^{num_pat}\s*(<=|<)\s*{var_pat}$", clause)
            if m3:
                val_str, op = m3.groups()
                val = float(val_str) if "." in val_str else int(val_str)
                rule["min"] = val
                matched = True

        # Pattern 4: N >= target or N > target (means target <= N)
        if not matched:
            m4 = re.match(rf"^{num_pat}\s*(>=|>)\s*{var_pat}$", clause)
            if m4:
                val_str, op = m4.groups()
                val = float(val_str) if "." in val_str else int(val_str)
                rule["max"] = val
                matched = True

        if not matched:
            unparsed_clauses.append(raw_clause)

    warning = None
    if unparsed_clauses:
        warning = (
            f"Constraint '{constraint_str}' contains logic that could not be "
            f"fully converted to standard min/max rules: "
            f"{', '.join(unparsed_clauses)}"
        )
    elif not rule and constraint_str:
        warning = (
            f"Constraint '{constraint_str}' could not be parsed into "
            "validation rules"
        )

    return rule, warning


def _parse_file_accept(body_accept: str) -> List[str]:
    """Extract list of allowed file extension strings from body::accept."""
    if not body_accept or not isinstance(body_accept, str):
        return []

    tokens = [t.strip().lower() for t in body_accept.split(",") if t.strip()]
    allowed: List[str] = []
    for token in tokens:
        if token == "image/*":
            continue
        clean_ext = token.lstrip(".")
        if clean_ext and clean_ext not in allowed:
            allowed.append(clean_ext)
    return allowed


def _extract_column_headers(
    sheet,
) -> Tuple[Dict[str, int], Dict[str, str], Dict[str, str]]:
    """Inspect row 1 of sheet to map column names to 0-based column indices.

    Returns:
      (col_map, label_langs, hint_langs)
      where label_langs maps col_name -> iso_code
    """
    col_map: Dict[str, int] = {}
    label_langs: Dict[str, str] = {}
    hint_langs: Dict[str, str] = {}

    for col_idx, cell in enumerate(sheet[1]):
        val = str(cell.value or "").strip()
        if not val:
            continue
        col_map[val.lower()] = col_idx

        val_lower = val.lower()
        if val_lower == "label":
            label_langs[val] = _DEFAULT_LANG_CODE
        elif val_lower.startswith("label::") or val_lower.startswith("label:"):
            lang_part = val.split(":", 2)[-1].strip()
            iso = _extract_iso(lang_part).strip().lower()
            label_langs[val] = iso or _DEFAULT_LANG_CODE

        if val_lower == "hint":
            hint_langs[val] = _DEFAULT_LANG_CODE
        elif val_lower.startswith("hint::") or val_lower.startswith("hint:"):
            lang_part = val.split(":", 2)[-1].strip()
            iso = _extract_iso(lang_part).strip().lower()
            hint_langs[val] = iso or _DEFAULT_LANG_CODE

    return col_map, label_langs, hint_langs


def _get_cell_value(
    row: Tuple, col_map: Dict[str, int], *keys: str
) -> Optional[str]:
    """Retrieve string value for the first matched key in row."""
    for key in keys:
        k_lower = key.lower()
        if k_lower in col_map:
            idx = col_map[k_lower]
            if idx < len(row):
                val = row[idx].value
                if val is not None:
                    s = str(val).strip()
                    if s:
                        return s
    return None


def parse_xlsform(file_or_stream: Any) -> Dict[str, Any]:
    """Parse an XLSForm (.xlsx) workbook into Akvo MIS form representation.

    Parameters:
      file_or_stream: File object, bytes, or file path to .xlsx workbook.

    Returns dict containing:
      form_name, version, default_language, languages, question_groups,
      total_questions, total_groups, skipped_rows, warnings, errors
    """
    try:
        wb = openpyxl.load_workbook(file_or_stream, data_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid Excel workbook: {exc}")

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
    if "survey" not in sheet_names_lower:
        raise ValueError("Missing required 'survey' sheet in XLSForm workbook")

    ws_survey = wb[sheet_names_lower["survey"]]
    ws_choices = (
        wb[sheet_names_lower["choices"]]
        if "choices" in sheet_names_lower
        else None
    )
    ws_settings = (
        wb[sheet_names_lower["settings"]]
        if "settings" in sheet_names_lower
        else None
    )

    # 1. Parse settings sheet
    form_name = "Imported Form"
    version = 1
    default_language = _DEFAULT_LANG_CODE
    all_languages: Set[str] = set()

    if ws_settings and ws_settings.max_row >= 2:
        s_col_map, _, _ = _extract_column_headers(ws_settings)
        row_2 = list(ws_settings.iter_rows(min_row=2, max_row=2))[0]
        title_val = _get_cell_value(row_2, s_col_map, "form_title", "title")
        if title_val:
            form_name = title_val
        version_val = _get_cell_value(row_2, s_col_map, "version")
        if version_val:
            try:
                version = int(float(version_val))
            except Exception:
                version = 1
        lang_val = _get_cell_value(row_2, s_col_map, "default_language")
        if lang_val:
            default_language = _extract_iso(lang_val).lower()
            all_languages.add(default_language)

    # 2. Parse choices sheet
    choices_map: Dict[str, List[Dict[str, Any]]] = {}
    if ws_choices and ws_choices.max_row >= 2:
        c_col_map, c_label_langs, _ = _extract_column_headers(ws_choices)
        for iso in c_label_langs.values():
            all_languages.add(iso)

        for row in ws_choices.iter_rows(min_row=2, values_only=False):
            list_name = _get_cell_value(
                row, c_col_map, "list_name", "list name"
            )
            opt_name = _get_cell_value(row, c_col_map, "name", "value")
            if not list_name or opt_name is None:
                continue

            primary_label = opt_name
            translations: List[Dict[str, str]] = []

            for col_header, iso in c_label_langs.items():
                val = _get_cell_value(row, c_col_map, col_header)
                if val is not None:
                    if iso == default_language and primary_label == opt_name:
                        primary_label = val
                    elif iso != default_language:
                        translations.append({"language": iso, "label": val})

            if primary_label == opt_name:
                for col_header, iso in c_label_langs.items():
                    val = _get_cell_value(row, c_col_map, col_header)
                    if val is not None:
                        primary_label = val
                        break

            opt_dict = {
                "order": len(choices_map.get(list_name, [])) + 1,
                "label": primary_label,
                "value": opt_name,
                "other": False,
                "translations": translations if translations else None,
            }
            choices_map.setdefault(list_name, []).append(opt_dict)

    # 3. Survey sheet columns and languages
    s_col_map, s_label_langs, s_hint_langs = _extract_column_headers(ws_survey)
    for iso in s_label_langs.values():
        all_languages.add(iso)
    for iso in s_hint_langs.values():
        all_languages.add(iso)

    if not all_languages:
        all_languages.add(_DEFAULT_LANG_CODE)
    if default_language not in all_languages:
        default_language = list(all_languages)[0]

    # PASS 1: Identify all survey rows, assign tmp_id, build name_to_tmp_id
    survey_rows_raw = list(ws_survey.iter_rows(min_row=2, values_only=False))
    name_to_tmp_id: Dict[str, int] = {}
    tmp_id_counter = 1
    duplicate_names: Set[str] = set()

    for row in survey_rows_raw:
        type_str = _get_cell_value(row, s_col_map, "type")
        if not type_str:
            continue
        type_clean = type_str.strip().lower()
        if type_clean.startswith("begin") or type_clean.startswith("end"):
            continue

        q_name = _get_cell_value(row, s_col_map, "name")
        if q_name:
            if q_name in name_to_tmp_id:
                duplicate_names.add(q_name)
            else:
                name_to_tmp_id[q_name] = tmp_id_counter
                tmp_id_counter += 1

    warnings: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    skipped_rows: List[Dict[str, Any]] = []

    if duplicate_names:
        for d_name in duplicate_names:
            errors.append(
                {
                    "path": f"question:{d_name}",
                    "message": (
                        f"Duplicate question name '{d_name}' in survey sheet"
                    ),
                    "level": "error",
                }
            )

    # PASS 2: Parse groups and questions
    question_groups: List[Dict[str, Any]] = []
    current_group: Optional[Dict[str, Any]] = None
    group_counter = 1
    total_questions = 0

    def _ensure_group(
        name_hint=None,
        label_hint=None,
        repeatable=False,
        repeat_text=None,
        translations=None,
    ):
        nonlocal current_group, group_counter
        if current_group is None:
            g_name = name_hint or f"group_{group_counter}"
            g_label = label_hint or f"Group {group_counter}"
            current_group = {
                "id": group_counter,
                "name": g_name,
                "label": g_label,
                "order": group_counter,
                "repeatable": repeatable,
                "repeat_text": repeat_text,
                "translations": translations,
                "question": [],
            }
            question_groups.append(current_group)
            group_counter += 1
        return current_group

    for row_idx, row in enumerate(survey_rows_raw, start=2):
        type_str = _get_cell_value(row, s_col_map, "type")
        if not type_str:
            continue

        type_clean = type_str.strip()
        type_lower = type_clean.lower()
        q_name = _get_cell_value(row, s_col_map, "name") or f"q_{row_idx}"
        appearance = _get_cell_value(row, s_col_map, "appearance")

        if type_lower in (
            "begin_group",
            "begin group",
            "begin_repeat",
            "begin repeat",
        ):
            is_repeat = type_lower.startswith(
                "begin_repeat"
            ) or type_lower.startswith("begin repeat")
            if current_group and not current_group["question"]:
                question_groups.remove(current_group)
                group_counter -= 1

            primary_label = _get_cell_value(row, s_col_map, "label") or q_name
            g_translations: List[Dict[str, str]] = []
            for col_header, iso in s_label_langs.items():
                val = _get_cell_value(row, s_col_map, col_header)
                if val is not None:
                    if iso == default_language:
                        primary_label = val
                    else:
                        g_translations.append(
                            {"language": iso, "name": val, "label": val}
                        )

            # Check repeat_count on begin_repeat
            repeat_count_val = _get_cell_value(row, s_col_map, "repeat_count")
            if is_repeat and repeat_count_val:
                warnings.append(
                    {
                        "path": f"row:{row_idx}",
                        "message": (
                            f"Repeat group '{q_name}' uses dynamic "
                            f"repeat_count '{repeat_count_val}'. In Akvo MIS, "
                            "repeatable groups allow manual entry "
                            "('Add another') without a fixed programmatic "
                            "count."
                        ),
                        "level": "warning",
                    }
                )

            # Check group-level relevant
            grp_relevant = _get_cell_value(row, s_col_map, "relevant")
            if grp_relevant:
                warnings.append(
                    {
                        "path": f"row:{row_idx}",
                        "message": (
                            f"Group '{q_name}' has a 'relevant' condition "
                            f"('{grp_relevant}'). Group-level relevance is "
                            "not natively supported in Akvo MIS; please "
                            "configure dependencies on individual questions "
                            "in the Form Editor."
                        ),
                        "level": "warning",
                    }
                )

            current_group = {
                "id": group_counter,
                "name": q_name,
                "label": primary_label,
                "order": group_counter,
                "repeatable": is_repeat,
                "repeat_text": "Add another" if is_repeat else None,
                "translations": g_translations if g_translations else None,
                "question": [],
            }
            question_groups.append(current_group)
            group_counter += 1
            continue

        if type_lower in (
            "end_group",
            "end group",
            "end_repeat",
            "end repeat",
        ):
            current_group = None
            continue

        akvo_type = None
        select_list_name = None
        has_or_other = False
        allow_decimal = False

        if type_lower == "text":
            akvo_type = "text"
        elif type_lower == "integer":
            akvo_type = "number"
            allow_decimal = False
        elif type_lower == "decimal":
            akvo_type = "number"
            allow_decimal = True
        elif type_lower == "date":
            akvo_type = "date"
        elif type_lower == "geopoint":
            akvo_type = "geo"
        elif type_lower == "image":
            if appearance and "signature" in appearance.lower():
                akvo_type = "signature"
            else:
                akvo_type = "image"
        elif type_lower == "file":
            akvo_type = "attachment"
        elif type_lower.startswith("select_one_from_file"):
            akvo_type = "cascade"
        elif type_lower.startswith("select_one ") or type_lower.startswith(
            "select_one\t"
        ):
            akvo_type = "option"
            parts = type_clean.split()
            if len(parts) >= 2:
                select_list_name = parts[1]
            if "or_other" in type_lower:
                has_or_other = True
        elif type_lower.startswith(
            "select_multiple "
        ) or type_lower.startswith("select_multiple\t"):
            akvo_type = "multiple_option"
            parts = type_clean.split()
            if len(parts) >= 2:
                select_list_name = parts[1]
            if "or_other" in type_lower:
                has_or_other = True
        elif type_lower == "calculate":
            calc_expr = _get_cell_value(row, s_col_map, "calculation")
            msg = f"Calculated field '{q_name}' (type: calculate"
            if calc_expr:
                msg += f", calculation: '{calc_expr}'"
            msg += ") is not supported in Akvo MIS — skipped"
            warn_obj = {
                "path": f"row:{row_idx}",
                "message": msg,
                "level": "warning",
            }
            skipped_rows.append(warn_obj)
            warnings.append(warn_obj)
            continue
        else:
            warn_obj = {
                "path": f"row:{row_idx}",
                "message": (
                    f"Unsupported XLSForm question type '{type_clean}' "
                    "— skipped"
                ),
                "level": "warning",
            }
            skipped_rows.append(warn_obj)
            warnings.append(warn_obj)
            continue

        grp = _ensure_group()

        primary_label = _get_cell_value(row, s_col_map, "label") or q_name
        primary_hint = _get_cell_value(row, s_col_map, "hint")

        q_translations_map: Dict[str, Dict[str, str]] = {}
        for col_header, iso in s_label_langs.items():
            val = _get_cell_value(row, s_col_map, col_header)
            if val is not None:
                if iso == default_language:
                    primary_label = val
                else:
                    q_translations_map.setdefault(iso, {})["label"] = val

        for col_header, iso in s_hint_langs.items():
            val = _get_cell_value(row, s_col_map, col_header)
            if val is not None:
                if iso == default_language:
                    primary_hint = val
                else:
                    q_translations_map.setdefault(iso, {})["tooltip"] = val

        q_translations = [
            {"language": iso, **data}
            for iso, data in q_translations_map.items()
            if data
        ]

        req_val = _get_cell_value(row, s_col_map, "required")
        is_required = bool(
            req_val and req_val.strip().lower() in ("yes", "true", "1")
        )

        rule_dict: Dict[str, Any] = {}
        if is_required:
            rule_dict["required"] = True
        if akvo_type == "number":
            rule_dict["allowDecimal"] = allow_decimal

        calc_expr = _get_cell_value(row, s_col_map, "calculation")
        if calc_expr:
            warnings.append(
                {
                    "path": f"row:{row_idx}",
                    "message": (
                        f"Question '{q_name}' has calculation '{calc_expr}'. "
                        "Dynamic calculations are not evaluated in Akvo MIS."
                    ),
                    "level": "warning",
                }
            )

        constraint_val = _get_cell_value(row, s_col_map, "constraint")
        if constraint_val:
            c_rule, c_warn = _parse_constraint(constraint_val, q_name)
            rule_dict.update(c_rule)
            if c_warn:
                warnings.append(
                    {
                        "path": f"row:{row_idx}",
                        "message": f"Question '{q_name}': {c_warn}",
                        "level": "warning",
                    }
                )

        if akvo_type == "attachment":
            body_accept = _get_cell_value(
                row, s_col_map, "body::accept", "body:accept", "accept"
            )
            if body_accept:
                allowed_types = _parse_file_accept(body_accept)
                if allowed_types:
                    rule_dict["allowedFileTypes"] = allowed_types

        q_options: Optional[List[Dict[str, Any]]] = None
        if select_list_name:
            base_opts = [
                dict(opt) for opt in choices_map.get(select_list_name, [])
            ]
            if has_or_other:
                base_opts.append(
                    {
                        "order": len(base_opts) + 1,
                        "label": "Other",
                        "value": "other",
                        "other": True,
                        "translations": None,
                    }
                )
            q_options = base_opts

        relevant_expr = _get_cell_value(row, s_col_map, "relevant")
        deps: List[Dict[str, Any]] = []
        dep_rule = "AND"
        if relevant_expr:
            parsed_deps, parsed_rule, dep_warn = parse_relevant_expression(
                relevant_expr, name_to_tmp_id
            )
            if dep_warn:
                warn_obj = {
                    "path": f"row:{row_idx}",
                    "message": (
                        f"Question '{q_name}' skip logic: {dep_warn} "
                        "— dependency skipped"
                    ),
                    "level": "warning",
                }
                warnings.append(warn_obj)
            else:
                deps = parsed_deps
                dep_rule = parsed_rule

        tmp_id = name_to_tmp_id.get(q_name) or total_questions + 1

        question_dict = {
            "id": tmp_id,
            "name": q_name,
            "label": primary_label,
            "type": akvo_type,
            "order": len(grp["question"]) + 1,
            "required": is_required,
            "tooltip": primary_hint,
            "rule": rule_dict if rule_dict else None,
            "dependency": deps if deps else None,
            "dependency_rule": dep_rule if deps else "AND",
            "option": q_options,
            "translations": q_translations if q_translations else None,
            "variable_name": q_name,
        }

        grp["question"].append(question_dict)
        total_questions += 1

    question_groups = [g for g in question_groups if g["question"]]

    return {
        "form_name": form_name,
        "version": version,
        "default_language": default_language,
        "languages": sorted(list(all_languages)),
        "question_groups": question_groups,
        "total_questions": total_questions,
        "total_groups": len(question_groups),
        "skipped_rows": skipped_rows,
        "warnings": warnings,
        "errors": errors,
    }


def validate_preflight(
    parsed: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Validate parsed XLSForm structure for preflight check.

    Returns:
      (errors, warnings)
    """
    errors: List[Dict[str, Any]] = list(parsed.get("errors", []))
    warnings: List[Dict[str, Any]] = list(parsed.get("warnings", []))

    if parsed.get("total_questions", 0) == 0:
        errors.append(
            {
                "path": "survey sheet",
                "message": "No valid questions found in survey sheet",
                "level": "error",
            }
        )

    return errors, warnings


def build_form_payload(
    parsed: Dict[str, Any],
    form_type: Any = "registration",
    parent_id: Optional[int] = None,
) -> Dict[str, Any]:
    """Convert parsed XLSForm structure into canonical form definition dict.

    Compatible with `normalize_form_definition()` and
    `import_form_definition()`.
    """
    if isinstance(form_type, str):
        type_val = (
            FormTypes.monitoring
            if form_type.lower() == "monitoring"
            else FormTypes.registration
        )
    else:
        type_val = form_type or FormTypes.registration

    parent_hint = None
    if parent_id is not None:
        try:
            parent_hint = {"id": int(parent_id)}
        except (ValueError, TypeError):
            parent_hint = None

    return {
        "_meta": None,
        "form_id": None,
        "name": parsed.get("form_name", "Imported Form"),
        "description": None,
        "type": type_val,
        "version": parsed.get("version") or 1,
        "languages": parsed.get("languages") or [_DEFAULT_LANG_CODE],
        "default_language": (
            parsed.get("default_language") or _DEFAULT_LANG_CODE
        ),
        "translations": None,
        "parent_hint": parent_hint,
        "question_group": parsed.get("question_groups", []),
    }
