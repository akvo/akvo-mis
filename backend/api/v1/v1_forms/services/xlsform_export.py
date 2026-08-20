import csv
import io
import re
from typing import Tuple, List, Dict, Any, Optional
from django.utils.translation import get_language_info
import openpyxl
from api.v1.v1_forms.constants import QuestionTypes
from api.v1.v1_profile.constants import DEFAULT_ADMINISTRATION_LEVELS
from api.v1.v1_profile.models import Administration, Levels

# XLSForm standard language names understood by KoboToolbox / pyxform.
# Maps ISO 639-1 codes to the 'Name (code)' format pyxform expects.
# KoboToolbox requires this format; bare codes (e.g. 'en') cause the
# 'unnamed translation' error in the form builder.
_LANG_NAMES: Dict[str, str] = {
    "en": "English (en)",
    "fr": "French (fr)",
    "es": "Spanish (es)",
    "ar": "Arabic (ar)",
    "pt": "Portuguese (pt)",
    "id": "Indonesian (id)",
    "de": "German (de)",
    "nl": "Dutch (nl)",
    "sw": "Swahili (sw)",
    "zh": "Chinese (zh)",
    "hi": "Hindi (hi)",
    "vi": "Vietnamese (vi)",
    "tet": "Tetum (tet)",
    "my": "Burmese (my)",
    "am": "Amharic (am)",
    "ja": "Japanese (ja)",
}

_DEFAULT_LANG_CODE = "en"
_DEFAULT_LANG_DISPLAY = _LANG_NAMES[_DEFAULT_LANG_CODE]

_XLSFORM_COLUMNS = [
    "type",
    "name",
    "label",
    "required",
    "hint",
    "relevant",
    "constraint",
    "constraint_message",
    "choice_filter",
    "parameters",
    "appearance",
    "calculation",
    "default",
    "repeat_count",
]


def _get_options_manager(question: Any) -> Optional[Any]:
    for attr in ("options", "option", "question_question_option"):
        opts = getattr(question, attr, None)
        if opts is not None:
            if hasattr(opts, "all"):
                return opts
            elif isinstance(opts, (list, tuple)):
                return _QuerySetAdapter(
                    [
                        _DictObject(o) if isinstance(o, dict) else o
                        for o in opts
                    ]
                )
    return None


def _map_type(question: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Maps Akvo MIS question type to (xlsform_type_str, appearance_or_None).
    Returns (None, None) for skipped types (tree, table, autofield).
    """
    qtype = question.type
    rule = getattr(question, "rule", None) or {}
    q_name = (
        getattr(question, "name", None)
        or f"q_{getattr(question, 'id', 'unknown')}"
    )

    if qtype in (QuestionTypes.text, QuestionTypes.input):
        return ("text", None)
    elif qtype == QuestionTypes.number:
        if rule.get("allowDecimal"):
            return ("decimal", None)
        return ("integer", None)
    elif qtype == QuestionTypes.date:
        return ("date", None)
    elif qtype == QuestionTypes.option:
        has_other = False
        opts_mgr = _get_options_manager(question)
        if opts_mgr:
            has_other = any(
                bool(getattr(opt, "other", False)) for opt in opts_mgr.all()
            )
        suffix = " or_other" if has_other else ""
        return (f"select_one option_{q_name}{suffix}", None)
    elif qtype == QuestionTypes.multiple_option:
        has_other = False
        opts_mgr = _get_options_manager(question)
        if opts_mgr:
            has_other = any(
                bool(getattr(opt, "other", False)) for opt in opts_mgr.all()
            )
        suffix = " or_other" if has_other else ""
        return (f"select_multiple option_{q_name}{suffix}", None)
    elif qtype == QuestionTypes.geo:
        return ("geopoint", None)
    elif qtype == QuestionTypes.geoshape:
        return ("geoshape", None)
    elif qtype == QuestionTypes.geotrace:
        return ("geotrace", None)
    elif qtype == QuestionTypes.image:
        return ("image", None)
    elif qtype == QuestionTypes.attachment:
        return ("file", None)
    elif qtype == QuestionTypes.signature:
        return ("image", "signature")
    elif qtype == QuestionTypes.cascade:
        return ("select_one_from_file administration.csv", None)
    elif qtype in (
        QuestionTypes.tree,
        QuestionTypes.table,
        QuestionTypes.autofield,
    ):
        return (None, None)

    return ("text", None)


def _get_levels_map(form: Any = None) -> Dict[int, str]:
    """
    Returns a mapping of {level_int: level_name}
    (e.g. {0: 'National', 1: 'Province', 2: 'District'}).
    Queries Levels model scoped to tenant if available,
    falling back to DEFAULT_ADMINISTRATION_LEVELS.
    """
    levels_map = {}
    try:
        tenant = getattr(form, "tenant", None)
        if tenant:
            qs = Levels.objects.filter(tenant=tenant).order_by("level")
        else:
            qs = Levels.objects.all().order_by("level")
        for lvl in qs:
            if lvl.level is not None and lvl.name:
                levels_map[lvl.level] = lvl.name
    except Exception:
        pass

    if not levels_map:
        try:
            for item in DEFAULT_ADMINISTRATION_LEVELS:
                lvl_idx = item.get("level")
                lvl_name = item.get("alias") or item.get("name")
                if lvl_idx is not None and lvl_name:
                    levels_map[lvl_idx] = lvl_name
        except Exception:
            pass

    return levels_map


def _get_cascade_levels(q: Any, levels_map: Dict[int, str]) -> List[int]:
    """
    Determines the list of level integers [min_level, ..., max_level]
    configured for a cascade question.
    By default, Level 0 (Country / National root) is omitted from survey
    questions so cascade selection begins at Level 1 (e.g. Region/Division),
    unless the hierarchy only has Level 0 or a positive min_level is set.
    """
    has_levels = bool(levels_map)
    default_min = (
        1 if (has_levels and 0 in levels_map and len(levels_map) > 1) else 0
    )
    min_level = default_min
    max_level = None

    q_api = getattr(q, "api", None)
    if q_api and isinstance(q_api, dict):
        if "min_level" in q_api and isinstance(q_api["min_level"], int):
            if q_api["min_level"] > 0:
                min_level = q_api["min_level"]
        if "max_level" in q_api and isinstance(q_api["max_level"], int):
            max_level = q_api["max_level"]

    if max_level is None:
        max_level = max(levels_map.keys()) if levels_map else 2

    if max_level < min_level:
        max_level = min_level

    return list(range(min_level, max_level + 1))


def _build_question_map(form: Any) -> Dict[int, Dict[str, Any]]:
    """
    Pre-pass building a map of {question_id: {"name": name, "type": type}}
    for dependency ID -> question name resolution.
    For multi-level cascade questions, maps to the deepest level question name.
    """
    levels_map = _get_levels_map(form)
    question_map = {}
    for group in form.form_question_group.all():
        for q in group.question_group_question.all():
            q_name = q.name or f"q_{q.id}"
            if q.type == QuestionTypes.cascade:
                levels = _get_cascade_levels(q, levels_map)
                if len(levels) > 1:
                    deepest_lvl = levels[-1]
                    q_name = f"{q_name}_level_{deepest_lvl}"
            question_map[q.id] = {
                "name": q_name,
                "type": q.type,
            }
    return question_map


def _lang_display(code: str) -> str:
    """
    Returns the full XLSForm language display name for a given ISO code
    using Django's built-in get_language_info (e.g. 'id' -> 'Indonesian (id)').
    Unknown codes get 'Code (code)' as fallback
    to ensure they are always named.
    KoboToolbox requires this format; bare codes cause 'unnamed translation'.
    """
    clean_code = (code or "").strip().lower()
    if not clean_code:
        return f"English ({_DEFAULT_LANG_CODE})"

    # Check manual overrides first if specified
    if clean_code in _LANG_NAMES:
        return _LANG_NAMES[clean_code]

    try:
        info = get_language_info(clean_code)
        name = (
            info.get("name")
            or info.get("name_local")
            or clean_code.capitalize()
        )
    except Exception:
        name = clean_code.capitalize()

    return f"{name} ({clean_code})"


def _extract_iso(display_or_code: str) -> str:
    """
    Extracts ISO code from 'English (en)' -> 'en' or returns code directly.
    """
    m = re.search(r"\(([^)]+)\)$", display_or_code)
    return m.group(1) if m else display_or_code


def _extract_translation(translations: Any, target_iso: str) -> Dict[str, str]:
    """
    Extracts translation dictionary for target ISO code.
    Handles list-of-dicts format used in DB:
    [{"language": "id", "name": "Judul", "label": "Judul", "tooltip": "Hint"}]
    and dict format:
        {"id": {"label": "Judul", "name": "Judul", "tooltip": "Hint"}}
    Returns dict with optional 'label' and 'hint' keys.
    """
    if not translations:
        return {}

    label_val: Optional[str] = None
    hint_val: Optional[str] = None

    if isinstance(translations, list):
        for item in translations:
            if isinstance(item, dict):
                lang = (
                    item.get("language")
                    or item.get("lang")
                    or item.get("code")
                )
                if lang == target_iso:
                    label_val = item.get("label") or item.get("name")
                    hint_val = (
                        item.get("tooltip")
                        or item.get("hint")
                        or item.get("text")
                    )
                    break
    elif isinstance(translations, dict):
        trans = translations.get(target_iso, {})
        if isinstance(trans, dict):
            label_val = trans.get("label") or trans.get("name")
            hint_val = (
                trans.get("tooltip") or trans.get("hint") or trans.get("text")
            )
            if isinstance(hint_val, dict):
                hint_val = hint_val.get("text") or hint_val.get("tooltip")

    res: Dict[str, str] = {}
    if label_val and isinstance(label_val, str) and label_val.strip():
        res["label"] = label_val.strip()
    if hint_val and isinstance(hint_val, str) and hint_val.strip():
        res["hint"] = hint_val.strip()
    return res


def _collect_translation_languages(form_obj: Any, codes: List[str]) -> None:
    """
    Helper to discover language codes defined inside
    form/group/question translations.
    """

    def _add_code(raw_code: Any) -> None:
        if raw_code and isinstance(raw_code, str) and raw_code.strip():
            c = raw_code.strip()
            if c not in codes:
                codes.append(c)

    # Form level
    f_trans = getattr(form_obj, "translations", None)
    if isinstance(f_trans, list):
        for t in f_trans:
            if isinstance(t, dict):
                _add_code(t.get("language") or t.get("lang") or t.get("code"))

    # Groups & questions level
    groups = getattr(form_obj, "form_question_group", None)
    if groups and hasattr(groups, "all"):
        for g in groups.all():
            g_trans = getattr(g, "translations", None)
            if isinstance(g_trans, list):
                for t in g_trans:
                    if isinstance(t, dict):
                        _add_code(
                            t.get("language") or t.get("lang") or t.get("code")
                        )
            qs = getattr(g, "question_group_question", None)
            if qs and hasattr(qs, "all"):
                for q in qs.all():
                    q_trans = getattr(q, "translations", None)
                    if isinstance(q_trans, list):
                        for t in q_trans:
                            if isinstance(t, dict):
                                _add_code(
                                    t.get("language")
                                    or t.get("lang")
                                    or t.get("code")
                                )


def _clean_lang_cols(
    raw_languages: Any,
    default_language: Optional[str] = None,
    form_obj: Optional[Any] = None,
) -> List[str]:
    """
    Returns list of XLSForm language display names
    (e.g. ['English (en)', 'Indonesian (id)']).
    Collects codes from default_language, raw_languages, and any translations
    found in form/groups/questions.
    """
    codes: List[str] = []

    # 1. Primary: default_language (if present)
    if (
        default_language
        and isinstance(default_language, str)
        and default_language.strip()
    ):
        codes.append(default_language.strip())

    # 2. Languages array
    if raw_languages and isinstance(raw_languages, list):
        for item in raw_languages:
            code = None
            if isinstance(item, str) and item.strip():
                code = item.strip()
            elif isinstance(item, dict):
                code = item.get("code") or item.get("name") or item.get("id")
            if code and isinstance(code, str) and code.strip():
                c_str = code.strip()
                if c_str not in codes:
                    codes.append(c_str)

    # 3. Scan form/group/question translations if form_obj provided
    if form_obj:
        _collect_translation_languages(form_obj, codes)

    if not codes:
        codes = [_DEFAULT_LANG_CODE]

    return [_lang_display(c) for c in codes]


def _build_settings_row(form: Any) -> Dict[str, Any]:
    """
    Builds the form metadata for the settings sheet.
    Always includes default_language in the named language display format
    (e.g. 'English (en)') so KoboToolbox can name the translation properly.
    """
    f_name = getattr(form, "name", None) or f"Form {getattr(form, 'id', '')}"
    lang_cols = _clean_lang_cols(
        getattr(form, "languages", None),
        getattr(form, "default_language", None),
        form_obj=form,
    )
    # lang_cols[0] is always the display name (e.g. 'English (en)')
    d_lang_display = lang_cols[0]
    return {
        "form_title": f_name,
        "form_id": f"form_{getattr(form, 'id', '')}",
        "version": str(getattr(form, "version", 1) or 1),
        "default_language": d_lang_display,
    }


def _build_choices_rows(
    form: Any, lang_cols: List[str]
) -> List[Dict[str, Any]]:
    """
    Builds list of choice dicts for select_one / select_multiple options.
    lang_cols contains display names like ['English (en)', 'French (fr)'].
    The first display name is used for the default language label.
    """
    choices = []
    d_lang_display = lang_cols[0]
    secondary_langs = [(disp, _extract_iso(disp)) for disp in lang_cols[1:]]

    for group in form.form_question_group.all():
        for q in group.question_group_question.all():
            if q.type in (QuestionTypes.option, QuestionTypes.multiple_option):
                list_name = f"option_{q.name or f'q_{q.id}'}"
                opts_mgr = _get_options_manager(q)
                options = opts_mgr.all() if opts_mgr else []
                for opt in options:
                    if opt.other:
                        continue
                        # 'or_other' handles the 'other'
                        # choice automatically in XLSForm
                    g_label = getattr(opt, "label", None) or str(opt.value)
                    row = {
                        "list_name": list_name,
                        "name": str(
                            opt.value if opt.value is not None else opt.id
                        ),
                        f"label::{d_lang_display}": g_label,
                    }

                    # Handle translations: secondary languages
                    # with primary fallback
                    for display, iso in secondary_langs:
                        trans_data = (
                            _extract_translation(opt.translations, iso)
                            if opt.translations
                            else {}
                        )
                        row[f"label::{display}"] = (
                            trans_data.get("label") or g_label
                        )
                    choices.append(row)
    return choices


def _build_relevant_expression(
    question: Any, question_map: Dict[int, Dict[str, Any]]
) -> str:
    """
    Converts question.dependency JSON list into an
    XLSForm XPath 'relevant' expression.
    Supported dependency conditions:
    - options (single or list) -> selected(${name}, 'val')
    - min -> ${name} >= N
    - max -> ${name} <= N
    - equal -> ${name} = 'val'
    - notEqual -> ${name} != 'val' and string-length(${name}) > 0
    Combined by dependency_rule ('AND' or 'OR', default 'AND').
    """
    deps = getattr(question, "dependency", None)
    if not deps or not isinstance(deps, list):
        return ""

    rule = (getattr(question, "dependency_rule", None) or "AND").upper()
    joiner = " or " if rule == "OR" else " and "

    expr_parts = []
    for dep in deps:
        if not isinstance(dep, dict):
            continue

        q_id = dep.get("id")
        if q_id not in question_map:
            continue

        q_target = question_map[q_id]
        target_name = q_target["name"]

        # 1. options condition
        if "options" in dep:
            opts = dep["options"]
            if isinstance(opts, list) and len(opts) > 0:
                if len(opts) == 1:
                    expr_parts.append(
                        f"selected(${{{target_name}}}, '{opts[0]}')"
                    )
                else:
                    sub = " or ".join(
                        f"selected(${{{target_name}}}, '{v}')" for v in opts
                    )
                    expr_parts.append(f"({sub})")

        # 2. min condition
        if "min" in dep:
            min_val = dep["min"]
            expr_parts.append(f"${{{target_name}}} >= {min_val}")

        # 3. max condition
        if "max" in dep:
            max_val = dep["max"]
            expr_parts.append(f"${{{target_name}}} <= {max_val}")

        # 4. equal condition
        if "equal" in dep:
            eq_val = dep["equal"]
            expr_parts.append(f"${{{target_name}}} = '{eq_val}'")

        # 5. notEqual condition
        if "notEqual" in dep:
            neq_val = dep["notEqual"]
            expr_parts.append(
                f"${{{target_name}}} != '{neq_val}' and string-length(${{{target_name}}}) > 0"  # noqa
            )

    return joiner.join(expr_parts)


def _build_constraint(rule: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Converts question rule dictionary into (constraint, constraint_message).
    Example:
    rule={"min": 1, "max": 7} ->
        (". >= 1 and . <= 7", "Value must be between 1 and 7")
    """
    if not rule or not isinstance(rule, dict):
        return None, None

    has_min = "min" in rule and rule["min"] is not None
    has_max = "max" in rule and rule["max"] is not None

    if has_min and has_max:
        min_val, max_val = rule["min"], rule["max"]
        return (
            f". >= {min_val} and . <= {max_val}",
            f"Value must be between {min_val} and {max_val}",
        )
    elif has_min:
        min_val = rule["min"]
        return f". >= {min_val}", f"Value must be at least {min_val}"
    elif has_max:
        max_val = rule["max"]
        return f". <= {max_val}", f"Value must be at most {max_val}"

    return None, None


def _build_survey_rows(
    form: Any, question_map: Dict[int, Dict[str, Any]], lang_cols: List[str]
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Builds flat list of survey sheet row dicts.
    lang_cols contains display names like ['English (en)', 'French (fr)'].
    Returns (survey_rows, skipped_question_names).
    """
    levels_map = _get_levels_map(form)
    survey_rows = []
    skipped = []
    d_lang_display = lang_cols[0]
    secondary_langs = [(disp, _extract_iso(disp)) for disp in lang_cols[1:]]

    for group in form.form_question_group.all():
        raw_g_name = group.name or f"group_{getattr(group, 'id', '1')}"
        child_names = [
            getattr(q, "name", None)
            for q in group.question_group_question.all()
            if getattr(q, "name", None)
        ]
        if raw_g_name in child_names:
            g_name = f"group_{raw_g_name}"
            skipped.append(
                f"group:{raw_g_name}->{g_name} (renamed group to avoid collision with child question)"  # noqa
            )
        else:
            g_name = raw_g_name
        g_is_repeat = bool(group.repeatable)

        # Emit begin_repeat / begin_group
        begin_type = "begin_repeat" if g_is_repeat else "begin_group"
        g_label = getattr(group, "label", None) or g_name
        begin_row = {
            "type": begin_type,
            "name": g_name,
            f"label::{d_lang_display}": g_label,
        }
        if g_is_repeat:
            l_q = getattr(group, "leading_question", None) or getattr(
                group, "leadingQuestion", None
            )
            if l_q:
                begin_row["repeat_count"] = f"count-selected(${{{l_q}}})"

        for display, iso in secondary_langs:
            trans_data = (
                _extract_translation(group.translations, iso)
                if group.translations
                else {}
            )
            begin_row[f"label::{display}"] = trans_data.get("label") or g_label
        survey_rows.append(begin_row)

        for q in group.question_group_question.all():
            q_name = q.name or f"q_{q.id}"
            xls_type, appearance = _map_type(q)

            if xls_type is None:
                skipped.append(q_name)
                continue

            relevant_expr = _build_relevant_expression(q, question_map)
            rule_obj = getattr(q, "rule", None)
            constraint_expr, constraint_msg = _build_constraint(rule_obj)

            hint_text = (
                q.tooltip.get("text")
                if (q.tooltip and isinstance(q.tooltip, dict))
                else None
            )
            primary_label = q.label or q_name

            if q.type == QuestionTypes.cascade:
                levels = _get_cascade_levels(q, levels_map)
                if len(levels) > 1:
                    for lvl_idx, lvl in enumerate(levels):
                        level_name = levels_map.get(lvl, f"Level {lvl}")
                        level_q_name = f"{q_name}_level_{lvl}"
                        level_primary_label = f"{primary_label} - {level_name}"

                        if lvl_idx == 0:
                            choice_filter = f"level = {lvl}"
                            level_relevant = relevant_expr
                        else:
                            prev_level_q_name = (
                                f"{q_name}_level_{levels[lvl_idx - 1]}"
                            )
                            choice_filter = (
                                f"parent_key = ${{{prev_level_q_name}}}"
                            )
                            parent_dep = f"${{{prev_level_q_name}}} != ''"
                            if relevant_expr:
                                level_relevant = (
                                    f"({relevant_expr}) and ({parent_dep})"
                                )
                            else:
                                level_relevant = parent_dep

                        level_row = {
                            "type": "select_one_from_file administration.csv",
                            "name": level_q_name,
                            "required": (
                                "yes"
                                if getattr(q, "required", False)
                                else "no"
                            ),
                            f"label::{d_lang_display}": level_primary_label,
                            "choice_filter": choice_filter,
                        }
                        if level_relevant:
                            level_row["relevant"] = level_relevant
                        if hint_text:
                            level_row[f"hint::{d_lang_display}"] = hint_text

                        for display, iso in secondary_langs:
                            trans_data = (
                                _extract_translation(q.translations, iso)
                                if q.translations
                                else {}
                            )
                            trans_base_label = (
                                trans_data.get("label") or primary_label
                            )
                            level_row[f"label::{display}"] = (
                                f"{trans_base_label} - {level_name}"
                            )
                            if hint_text:
                                level_row[f"hint::{display}"] = (
                                    trans_data.get("hint") or hint_text
                                )
                            elif "hint" in trans_data:
                                level_row[f"hint::{display}"] = trans_data[
                                    "hint"
                                ]

                        survey_rows.append(level_row)
                    continue
                else:
                    # Single level cascade
                    lvl = levels[0]
                    single_choice_filter = f"level = {lvl}"
            else:
                single_choice_filter = None

            q_row = {
                "type": xls_type,
                "name": q_name,
                "required": "yes" if getattr(q, "required", False) else "no",
                f"label::{d_lang_display}": primary_label,
            }

            if single_choice_filter:
                q_row["choice_filter"] = single_choice_filter
            if relevant_expr:
                q_row["relevant"] = relevant_expr
            if constraint_expr:
                q_row["constraint"] = constraint_expr
                q_row["constraint_message"] = constraint_msg
            if appearance:
                q_row["appearance"] = appearance

            if hint_text:
                q_row[f"hint::{d_lang_display}"] = hint_text

            primary_hint = q_row.get(f"hint::{d_lang_display}")
            for display, iso in secondary_langs:
                trans_data = (
                    _extract_translation(q.translations, iso)
                    if q.translations
                    else {}
                )
                q_row[f"label::{display}"] = (
                    trans_data.get("label") or primary_label
                )
                if primary_hint:
                    q_row[f"hint::{display}"] = (
                        trans_data.get("hint") or primary_hint
                    )
                elif "hint" in trans_data:
                    q_row[f"hint::{display}"] = trans_data["hint"]

            survey_rows.append(q_row)

        # Emit end_repeat / end_group
        end_type = "end_repeat" if g_is_repeat else "end_group"
        survey_rows.append({"type": end_type})

    return survey_rows, skipped


class _DictObject:
    def __init__(self, d: dict):
        self._d = d
        self.repeatable = False
        self.label = None
        self.tooltip = None
        self.translations = None
        self.rule = None
        self.dependency = None
        self.dependency_rule = None
        self.required = False
        self.other = False
        self.api = None
        for k, v in d.items():
            setattr(self, k, v)
        if "defaultLanguage" in d and not hasattr(self, "default_language"):
            self.default_language = d["defaultLanguage"]
        if "shortLabel" in d and not hasattr(self, "short_label"):
            self.short_label = d["shortLabel"]
        if "dependencyRule" in d:
            self.dependency_rule = d["dependencyRule"]


class _QuerySetAdapter:
    def __init__(self, items: list):
        self._items = items

    def all(self):
        return self._items


def _adapt_form_dict(data: dict) -> Any:
    """
    Wraps a form payload dict (from published version snapshot or serializer)
    into an object exposing the Django ORM model interface
    expected by the generator.
    """
    groups = []
    raw_groups = (
        data.get("question_group") or data.get("question_groups") or []
    )
    for g_dict in raw_groups:
        questions = []
        raw_questions = g_dict.get("question") or g_dict.get("questions") or []
        for q_dict in raw_questions:
            q_obj = _DictObject(q_dict)
            if "repeat" in q_dict:
                q_obj.repeatable = bool(q_dict["repeat"])
            if "rule" in q_dict and isinstance(q_dict["rule"], dict):
                q_obj.required = bool(q_dict["rule"].get("required", False))

            # Map type string (e.g. 'text', 'cascade', 'attachment')
            # to QuestionTypes constant int if necessary
            if isinstance(q_obj.type, str):
                type_name = q_obj.type.lower()
                if hasattr(QuestionTypes, type_name):
                    q_obj.type = getattr(QuestionTypes, type_name)
                elif type_name == "input":
                    q_obj.type = QuestionTypes.input
                elif type_name == "multiple_option":
                    q_obj.type = QuestionTypes.multiple_option
                elif type_name == "cascade":
                    q_obj.type = QuestionTypes.cascade
                elif type_name == "attachment":
                    q_obj.type = QuestionTypes.attachment
                elif type_name == "signature":
                    q_obj.type = QuestionTypes.signature
                elif type_name == "geoshape":
                    q_obj.type = QuestionTypes.geoshape
                elif type_name == "geotrace":
                    q_obj.type = QuestionTypes.geotrace
                elif type_name == "autofield":
                    q_obj.type = QuestionTypes.autofield
                elif type_name == "tree":
                    q_obj.type = QuestionTypes.tree
                elif type_name == "table":
                    q_obj.type = QuestionTypes.table
                else:
                    q_obj.type = QuestionTypes.text

            # Wrap options
            raw_options = q_dict.get("options") or q_dict.get("option") or []
            q_obj.options = _QuerySetAdapter(
                [
                    _DictObject(o) if isinstance(o, dict) else o
                    for o in raw_options
                ]
            )

            questions.append(q_obj)

        g_obj = _DictObject(g_dict)
        if "repeat" in g_dict:
            g_obj.repeatable = bool(g_dict["repeat"])
        g_obj.question_group_question = _QuerySetAdapter(questions)
        groups.append(g_obj)

    form_obj = _DictObject(data)
    form_obj.form_question_group = _QuerySetAdapter(groups)
    return form_obj


def generate_xlsform(form: Any) -> Tuple[io.BytesIO, List[str]]:
    """
    Public entrypoint: generates XLSForm workbook for a given form.
    Supports both Forms Django ORM model instances and dictionary
    payloads (snapshots).
    """
    if isinstance(form, dict):
        form = _adapt_form_dict(form)

    wb = openpyxl.Workbook()
    ws_survey = wb.active
    ws_survey.title = "survey"
    ws_choices = wb.create_sheet(title="choices")
    ws_settings = wb.create_sheet(title="settings")

    lang_cols = _clean_lang_cols(
        getattr(form, "languages", None),
        getattr(form, "default_language", None),
    )
    question_map = _build_question_map(form)

    # 1. Settings Sheet
    settings_data = _build_settings_row(form)
    ws_settings.append(list(settings_data.keys()))
    ws_settings.append(list(settings_data.values()))

    # 2. Survey Sheet
    survey_rows, skipped = _build_survey_rows(form, question_map, lang_cols)

    # Build header columns — always language-tagged
    cols = []
    for c in _XLSFORM_COLUMNS:
        if c == "label":
            for display in lang_cols:
                cols.append(f"label::{display}")
        elif c == "hint":
            for display in lang_cols:
                cols.append(f"hint::{display}")
        else:
            cols.append(c)

    seen: set = set()
    final_cols: List[str] = []
    for c in cols:
        if c not in seen:
            seen.add(c)
            final_cols.append(c)
    cols = final_cols

    ws_survey.append(cols)
    for row in survey_rows:
        ws_survey.append([row.get(col, "") for col in cols])

    # 3. Choices Sheet — always language-tagged
    choices_rows = _build_choices_rows(form, lang_cols)
    choice_cols = ["list_name", "name"] + [
        f"label::{display}" for display in lang_cols
    ]
    ws_choices.append(choice_cols)
    for row in choices_rows:
        ws_choices.append([row.get(col, "") for col in choice_cols])

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, skipped


def generate_administration_csv(form: Any, user: Any) -> str:
    """
    Generates a lookup CSV stream for cascade questions in XLSForm format.
    Columns: list_name, name, label, parent_key, level
    - Filters by Administration.objects.for_user(user)
    - If form has cascade questions with 'max_level' specified in question.api,
      caps the exported levels to max(max_level).
    """
    if isinstance(form, dict):
        form = _adapt_form_dict(form)

    max_level = None
    for group in form.form_question_group.all():
        for q in group.question_group_question.all():
            if (
                q.type == QuestionTypes.cascade
                and q.api
                and isinstance(q.api, dict)
            ):
                lvl = q.api.get("max_level")
                if lvl is not None and isinstance(lvl, int):
                    max_level = (
                        max(max_level, lvl) if max_level is not None else lvl
                    )

    qs = Administration.objects.for_user(user).select_related(
        "level", "parent"
    )
    if max_level is not None:
        qs = qs.filter(level__level__lte=max_level)

    qs = qs.order_by("level__level", "id")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["list_name", "name", "label", "parent_key", "level"])

    for adm in qs:
        name_val = adm.code if adm.code else str(adm.id)
        label_val = adm.name or ""
        parent_key_val = ""
        if adm.parent:
            parent_key_val = (
                adm.parent.code if adm.parent.code else str(adm.parent.id)
            )
        lvl_val = (
            str(adm.level.level)
            if adm.level and adm.level.level is not None
            else "0"
        )
        writer.writerow(
            ["administration", name_val, label_val, parent_key_val, lvl_val]
        )

    return output.getvalue()
