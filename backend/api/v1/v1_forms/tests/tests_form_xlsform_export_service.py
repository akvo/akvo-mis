import openpyxl
from unittest.mock import MagicMock

from django.test import TestCase

from api.v1.v1_forms.constants import QuestionTypes
from api.v1.v1_forms.services.xlsform_export import (
    _map_type,
    _build_question_map,
    _build_settings_row,
    _build_choices_rows,
    _build_survey_rows,
    _build_relevant_expression,
    _lang_display,
    _extract_iso,
    generate_xlsform,
)


class DummyObject:
    def __init__(self, **kwargs):
        self.label = None
        self.rule = None
        self.tooltip = None
        self.translations = None
        self.api = None
        self.repeatable = False
        self.required = False
        for k, v in kwargs.items():
            setattr(self, k, v)
            if k == "repeat":
                self.repeatable = bool(v)
            if k == "rule" and isinstance(v, dict) and "required" in v:
                self.required = bool(v["required"])


class XLSFormExportServiceTestCase(TestCase):
    def test_map_type_all_17_types(self):
        # 1. Input / Text
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.text)), ("text", None)
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.input)), ("text", None)
        )

        # 2. Number
        self.assertEqual(
            _map_type(
                DummyObject(
                    type=QuestionTypes.number, rule={"allowDecimal": True}
                )
            ),
            ("decimal", None),
        )
        self.assertEqual(
            _map_type(
                DummyObject(
                    type=QuestionTypes.number, rule={"allowDecimal": False}
                )
            ),
            ("integer", None),
        )

        # 3. Date
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.date)), ("date", None)
        )

        # 4. Option (Single & Multiple with or_other)
        opt_normal = DummyObject(other=False)
        opt_other = DummyObject(other=True)

        mock_mgr_normal = MagicMock()
        mock_mgr_normal.all.return_value = [opt_normal]

        mock_mgr_other = MagicMock()
        mock_mgr_other.all.return_value = [opt_normal, opt_other]

        q_opt = DummyObject(
            type=QuestionTypes.option,
            name="water_source",
            options=mock_mgr_normal,
        )
        self.assertEqual(
            _map_type(q_opt), ("select_one option_water_source", None)
        )

        q_opt_other = DummyObject(
            type=QuestionTypes.option,
            name="water_source",
            options=mock_mgr_other,
        )
        self.assertEqual(
            _map_type(q_opt_other),
            ("select_one option_water_source or_other", None),
        )

        q_mopt = DummyObject(
            type=QuestionTypes.multiple_option,
            name="amenities",
            options=mock_mgr_normal,
        )
        self.assertEqual(
            _map_type(q_mopt), ("select_multiple option_amenities", None)
        )

        q_mopt_other = DummyObject(
            type=QuestionTypes.multiple_option,
            name="amenities",
            options=mock_mgr_other,
        )
        self.assertEqual(
            _map_type(q_mopt_other),
            ("select_multiple option_amenities or_other", None),
        )

        # 5. Geo type
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.geo)), ("geopoint", None)
        )

        # 6. Media & File
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.image)), ("image", None)
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.attachment)),
            ("file", None),
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.signature)),
            ("image", "signature"),
        )

        # 7. Cascade
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.cascade)),
            ("select_one_from_file administration.csv", None),
        )

        # 8. Skipped types (tree, table, autofield, geoshape, geotrace)
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.tree)), (None, None)
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.table)), (None, None)
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.autofield)), (None, None)
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.geoshape)), (None, None)
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.geotrace)), (None, None)
        )

        # 9. Fallback for unknown
        self.assertEqual(_map_type(DummyObject(type=999)), ("text", None))

    def test_build_question_map(self):
        q1 = DummyObject(id=1, name="q_one", type=QuestionTypes.text)
        q2 = DummyObject(id=2, name=None, type=QuestionTypes.number)
        g = DummyObject(question_group_question=MagicMock())
        g.question_group_question.all.return_value = [q1, q2]

        f = DummyObject(form_question_group=MagicMock())
        f.form_question_group.all.return_value = [g]

        qmap = _build_question_map(f)
        self.assertEqual(
            qmap,
            {
                1: {"name": "q_one", "type": QuestionTypes.text},
                2: {"name": "q_2", "type": QuestionTypes.number},
            },
        )

    def test_build_settings_row(self):
        f = DummyObject(
            id=42,
            name="Test Form",
            version=2,
            default_language="fr",
            languages=["fr", "en"],
        )
        settings = _build_settings_row(f)
        self.assertEqual(
            settings,
            {
                "form_title": "Test Form",
                "form_id": "form_42",
                "version": "2",
                "default_language": "French (fr)",
            },
        )

    def test_build_choices_rows_with_translations_and_other_filter(self):
        opt1 = DummyObject(
            id=10,
            value="yes",
            label="Yes",
            other=False,
            translations={"en": {"label": "Yes (EN)"}},
        )
        opt2 = DummyObject(
            id=11,
            value="other_val",
            label="Other",
            other=True,  # should be skipped
            translations=None,
        )

        mock_opts = MagicMock()
        mock_opts.all.return_value = [opt1, opt2]

        q = DummyObject(
            type=QuestionTypes.option,
            name="consent",
            options=mock_opts,
        )
        g = DummyObject(question_group_question=MagicMock())
        g.question_group_question.all.return_value = [q]

        f = DummyObject(form_question_group=MagicMock())
        f.form_question_group.all.return_value = [g]

        # Two languages: primary French (fr), secondary English (en)
        choices = _build_choices_rows(f, ["French (fr)", "English (en)"])
        self.assertEqual(len(choices), 1)
        self.assertEqual(
            choices[0],
            {
                "list_name": "option_consent",
                "name": "yes",
                "label::French (fr)": "Yes",  # opt.label (default)
                "label::English (en)": "Yes (EN)",  # from translations
            },
        )

    def test_build_survey_rows_repeatable_group_and_multilingual_hints(self):
        q_text = DummyObject(
            id=1,
            name="member_name",
            type=QuestionTypes.text,
            label="Member Name",
            tooltip={"text": "Enter full name"},
            rule={"required": True},
            translations={
                "en": {"label": "Member Name EN", "tooltip": "Full name EN"}
            },
        )

        g_repeat = DummyObject(
            name="household_members",
            repeat=True,
            question_group_question=MagicMock(),
        )
        g_repeat.question_group_question.all.return_value = [q_text]

        f = DummyObject(form_question_group=MagicMock())
        f.form_question_group.all.return_value = [g_repeat]

        q_map = {1: {"name": "member_name", "type": QuestionTypes.text}}
        # Two display-name langs: primary Spanish, secondary English
        rows, skipped = _build_survey_rows(
            f, q_map, ["Spanish (es)", "English (en)"]
        )

        self.assertEqual(len(skipped), 0)
        self.assertEqual(len(rows), 3)  # begin_repeat, question, end_repeat

        # 1. begin_repeat row
        self.assertEqual(rows[0]["type"], "begin_repeat")
        self.assertEqual(rows[0]["name"], "household_members")
        # Group has no label field — falls back to name
        self.assertIn("label::Spanish (es)", rows[0])

        # 2. question row
        q_row = rows[1]
        self.assertEqual(q_row["type"], "text")
        self.assertEqual(q_row["name"], "member_name")
        self.assertEqual(q_row["required"], "yes")
        # Default lang label
        self.assertEqual(q_row["label::Spanish (es)"], "Member Name")
        # Hint for default lang (from tooltip)
        self.assertEqual(q_row["hint::Spanish (es)"], "Enter full name")
        # Translation for English (en) ISO code 'en'
        self.assertEqual(q_row["label::English (en)"], "Member Name EN")
        self.assertEqual(q_row["hint::English (en)"], "Full name EN")

        # 3. end_repeat row
        self.assertEqual(rows[2]["type"], "end_repeat")

    def test_generate_xlsform_full_excel_verification(self):
        q1 = DummyObject(
            id=1,
            name="q_text",
            type=QuestionTypes.text,
            label="Text Question",
            rule={},
            translations=None,
            tooltip=None,
        )

        g = DummyObject(
            name="group_one", repeat=False, question_group_question=MagicMock()
        )
        g.question_group_question.all.return_value = [q1]

        # No explicit languages — should default to English (en)
        f = DummyObject(
            id=10,
            name="Test Form",
            version=1,
            default_language=None,
            languages=None,
            form_question_group=MagicMock(),
        )
        f.form_question_group.all.return_value = [g]

        stream, skipped = generate_xlsform(f)
        self.assertEqual(skipped, [])

        # Load generated workbook from stream
        wb = openpyxl.load_workbook(stream)
        self.assertEqual(wb.sheetnames, ["survey", "choices", "settings"])

        ws_survey = wb["survey"]
        ws_settings = wb["settings"]

        # Check headers of survey sheet — must use named language format
        headers = [cell.value for cell in ws_survey[1]]
        self.assertIn("type", headers)
        self.assertIn("name", headers)
        self.assertIn("label::English (en)", headers)
        self.assertNotIn("label", headers)  # bare label must not appear
        self.assertNotIn("label::en", headers)  # bare code must not appear

        # Check settings — default_language must be in display format
        settings_headers = [cell.value for cell in ws_settings[1]]
        settings_values = [cell.value for cell in ws_settings[2]]
        s_map = dict(zip(settings_headers, settings_values))
        self.assertEqual(s_map["form_title"], "Test Form")
        self.assertEqual(s_map["default_language"], "English (en)")

    def test_generate_xlsform_dict_payload_with_indonesian_translations(self):
        dict_payload = {
            "id": 4,
            "name": "Test Form 4",
            "version": 1,
            "languages": ["en", "id"],
            "defaultLanguage": "en",
            "translations": [
                {"language": "id", "name": "Formulir Percobaan 4"}
            ],
            "question_group": [
                {
                    "id": 44,
                    "name": "completeness_check",
                    "label": "Completeness Check",
                    "translations": [
                        {"language": "id", "name": "Memastikan Komplit"}
                    ],
                    "question": [
                        {
                            "id": 442,
                            "name": "name",
                            "label": "Your full name",
                            "type": "text",
                            "required": True,
                            "translations": [
                                {"language": "id", "name": "Nama lengkap"}
                            ],
                        }
                    ],
                }
            ],
        }

        stream, skipped = generate_xlsform(dict_payload)
        self.assertEqual(skipped, [])

        wb = openpyxl.load_workbook(stream)
        ws_survey = wb["survey"]

        headers = [cell.value for cell in ws_survey[1]]
        self.assertIn("label::English (en)", headers)
        self.assertIn("label::Indonesian (id)", headers)

        # Check row 2 (group)
        r2 = [cell.value for cell in ws_survey[2]]
        r2_map = dict(zip(headers, r2))
        self.assertEqual(r2_map["label::English (en)"], "Completeness Check")
        self.assertEqual(
            r2_map["label::Indonesian (id)"], "Memastikan Komplit"
        )

        # Check row 3 (question 'name')
        r3 = [cell.value for cell in ws_survey[3]]
        r3_map = dict(zip(headers, r3))
        self.assertEqual(r3_map["label::English (en)"], "Your full name")
        self.assertEqual(r3_map["label::Indonesian (id)"], "Nama lengkap")

    def test_missing_translation_falls_back_to_primary_label(self):
        dict_payload = {
            "id": 5,
            "name": "Fallback Test Form",
            "version": 1,
            "languages": ["en", "id"],
            "defaultLanguage": "en",
            "question_group": [
                {
                    "id": 10,
                    "name": "g1",
                    "label": "Group One",
                    "question": [
                        {
                            "id": 101,
                            "name": "q_no_trans",
                            "label": "Untranslated Question",
                            "type": "text",
                            "translations": None,
                        }
                    ],
                }
            ],
        }
        stream, _ = generate_xlsform(dict_payload)
        wb = openpyxl.load_workbook(stream)
        ws_survey = wb["survey"]
        headers = [cell.value for cell in ws_survey[1]]
        row_q = dict(zip(headers, [cell.value for cell in ws_survey[3]]))

        # Secondary lang (id) must fall back to primary label instead of empty
        self.assertEqual(row_q["label::English (en)"], "Untranslated Question")
        self.assertEqual(
            row_q["label::Indonesian (id)"], "Untranslated Question"
        )

    def test_lang_display_edge_cases(self):
        # 1. Standard mapped codes
        self.assertEqual(_lang_display("en"), "English (en)")
        self.assertEqual(_lang_display("id"), "Indonesian (id)")
        self.assertEqual(_lang_display("fr"), "French (fr)")

        # 2. ISO 639 codes resolved via Django get_language_info
        self.assertEqual(_lang_display("tet"), "Tetum (tet)")
        self.assertEqual(_lang_display("ja"), "Japanese (ja)")
        self.assertEqual(_lang_display("am"), "Amharic (am)")

        # 3. Custom / unmapped ISO code fallback
        self.assertEqual(_lang_display("xyz"), "Xyz (xyz)")

        # 4. Whitespace and uppercase handling
        self.assertEqual(_lang_display("  ID  "), "Indonesian (id)")
        self.assertEqual(_lang_display("TET"), "Tetum (tet)")

        # 5. Empty / None fallback
        self.assertEqual(_lang_display(""), "English (en)")
        self.assertEqual(_lang_display(None), "English (en)")

        # 6. Extract ISO matching
        self.assertEqual(_extract_iso("Indonesian (id)"), "id")
        self.assertEqual(_extract_iso("Tetum (tet)"), "tet")
        self.assertEqual(_extract_iso("Xyz (xyz)"), "xyz")
        self.assertEqual(_extract_iso("raw_code"), "raw_code")

    def test_export_form_with_unmapped_language_code(self):
        dict_payload = {
            "id": 88,
            "name": "Unmapped Language Form",
            "version": 1,
            "languages": ["ko", "xyz"],
            "defaultLanguage": "ko",
            "question_group": [
                {
                    "id": 1,
                    "name": "g1",
                    "label": "Korean Group",
                    "question": [
                        {
                            "id": 10,
                            "name": "q1",
                            "label": "Korean Question",
                            "type": "text",
                            "translations": [
                                {
                                    "language": "xyz",
                                    "label": "Custom Lang Question",
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        stream, _ = generate_xlsform(dict_payload)
        wb = openpyxl.load_workbook(stream)

        ws_settings = wb["settings"]
        self.assertEqual(
            ws_settings.cell(row=2, column=4).value, "Korean (ko)"
        )

        ws_survey = wb["survey"]
        headers = [cell.value for cell in ws_survey[1]]
        self.assertIn("label::Korean (ko)", headers)
        self.assertIn("label::Xyz (xyz)", headers)

        row_q = dict(zip(headers, [cell.value for cell in ws_survey[3]]))
        self.assertEqual(row_q["label::Korean (ko)"], "Korean Question")
        self.assertEqual(row_q["label::Xyz (xyz)"], "Custom Lang Question")

    def test_empty_form_export(self):
        empty_payload = {
            "id": 99,
            "name": "Empty Form",
            "version": 1,
            "question_group": [],
        }
        stream, skipped = generate_xlsform(empty_payload)
        self.assertIsNotNone(stream)
        self.assertEqual(skipped, [])
        wb = openpyxl.load_workbook(stream)
        self.assertIn("survey", wb.sheetnames)
        self.assertIn("choices", wb.sheetnames)
        self.assertIn("settings", wb.sheetnames)

    def test_dangling_dependency_reference(self):
        q_with_deleted_dep = DummyObject(
            name="q2",
            dependency=[
                {"id": 99999, "options": ["yes"]},  # Question 99999 deleted
                {"id": 100, "min": 5},
            ],
            dependency_rule="AND",
        )
        q_map = {100: {"name": "existing_q", "type": QuestionTypes.number}}
        expr = _build_relevant_expression(q_with_deleted_dep, q_map)
        # Dangling ID 99999 is skipped cleanly; only valid ID 100 is emitted
        self.assertEqual(expr, "${existing_q} >= 5")

    def test_skipped_question_types_warning_header(self):
        payload_with_skipped = {
            "id": 77,
            "name": "Form With Webform Types",
            "version": 1,
            "question_group": [
                {
                    "id": 1,
                    "name": "g1",
                    "label": "Group 1",
                    "question": [
                        {
                            "id": 10,
                            "name": "tree_q",
                            "label": "Tree Q",
                            "type": "tree",
                        },
                        {
                            "id": 11,
                            "name": "table_q",
                            "label": "Table Q",
                            "type": "table",
                        },
                        {
                            "id": 12,
                            "name": "auto_q",
                            "label": "Auto Q",
                            "type": "autofield",
                        },
                    ],
                }
            ],
        }
        stream, skipped = generate_xlsform(payload_with_skipped)
        self.assertIn("tree_q", skipped)
        self.assertIn("table_q", skipped)
        self.assertIn("auto_q", skipped)

    def test_repeat_group_leading_question(self):
        payload_repeat_leading = {
            "id": 55,
            "name": "Repeat Leading Form",
            "version": 1,
            "question_group": [
                {
                    "id": 1,
                    "name": "household_members",
                    "label": "Household Members",
                    "repeatable": True,
                    "leading_question": "num_members",
                    "question": [
                        {
                            "id": 10,
                            "name": "member_name",
                            "label": "Member Name",
                            "type": "text",
                        }
                    ],
                }
            ],
        }
        stream, _ = generate_xlsform(payload_repeat_leading)
        wb = openpyxl.load_workbook(stream)
        ws_survey = wb["survey"]
        headers = [cell.value for cell in ws_survey[1]]
        row_begin_repeat = dict(
            zip(headers, [cell.value for cell in ws_survey[2]])
        )
        self.assertEqual(row_begin_repeat["type"], "begin_repeat")
        self.assertEqual(
            row_begin_repeat["repeat_count"], "count-selected(${num_members})"
        )

    def test_group_and_question_name_collision(self):
        payload_collision = {
            "id": 55,
            "name": "Collision Form",
            "version": 1,
            "question_group": [
                {
                    "id": 1,
                    "name": "signature",
                    "label": "Signature Group",
                    "question": [
                        {
                            "id": 10,
                            "name": "signature",
                            "label": "Signature Question",
                            "type": "signature",
                        }
                    ],
                }
            ],
        }
        stream, skipped = generate_xlsform(payload_collision)
        self.assertIn(
            "group:signature->group_signature (renamed group to avoid collision with child question)",  # noqa
            skipped,
        )
        wb = openpyxl.load_workbook(stream)
        ws_survey = wb["survey"]
        headers = [cell.value for cell in ws_survey[1]]
        row_group = dict(zip(headers, [cell.value for cell in ws_survey[2]]))
        row_q = dict(zip(headers, [cell.value for cell in ws_survey[3]]))

        # Group name prefixed to 'group_signature',
        # Question name remains 'signature'
        self.assertEqual(row_group["type"], "begin_group")
        self.assertEqual(row_group["name"], "group_signature")
        self.assertEqual(row_q["type"], "image")
        self.assertEqual(row_q["name"], "signature")

    def test_snapshot_dict_with_option_key_generates_choices_sheet(self):
        payload_snapshot = {
            "id": 6001,
            "name": "Visualization Test Registration",
            "version": 2,
            "question_group": [
                {
                    "id": 1,
                    "name": "registration_info",
                    "label": "Registration Info",
                    "question": [
                        {
                            "id": 101,
                            "name": "site_type",
                            "label": "Site Type",
                            "type": "option",
                            "option": [
                                {"id": 1, "value": "urban", "label": "Urban"},
                                {"id": 2, "value": "rural", "label": "Rural"},
                            ],
                        }
                    ],
                }
            ],
        }
        stream, _ = generate_xlsform(payload_snapshot)
        wb = openpyxl.load_workbook(stream)
        ws_choices = wb["choices"]
        rows = list(ws_choices.iter_rows(values_only=True))
        self.assertGreaterEqual(len(rows), 3)  # header + 2 options
        self.assertEqual(rows[0], ("list_name", "name", "label::English (en)"))
        self.assertEqual(rows[1], ("option_site_type", "urban", "Urban"))
        self.assertEqual(rows[2], ("option_site_type", "rural", "Rural"))

    def test_cascade_expansion_omits_level_zero_and_adds_progressive_relevance(  # noqa
        self,
    ):
        payload_cascade = {
            "id": 9001,
            "name": "Cascade Relevance Form",
            "version": 1,
            "question_group": [
                {
                    "id": 1,
                    "name": "location_group",
                    "label": "Location Group",
                    "question": [
                        {
                            "id": 501,
                            "name": "location",
                            "label": "Farm Location",
                            "type": "cascade",
                        }
                    ],
                }
            ],
        }
        stream, _ = generate_xlsform(payload_cascade)
        wb = openpyxl.load_workbook(stream)
        ws_survey = wb["survey"]
        headers = [cell.value for cell in ws_survey[1]]
        name_idx = headers.index("name")
        type_idx = headers.index("type")
        filter_idx = headers.index("choice_filter")
        relevant_idx = headers.index("relevant")
        label_idx = headers.index("label::English (en)")

        rows = list(ws_survey.iter_rows(values_only=True))[1:]  # skip header
        # Row 0: begin_group
        # Row 1: location_level_1 (Region/Province)
        # Row 2: location_level_2 (District)
        # Row 3: location_level_3 (Subdistrict)
        # Row 4: location_level_4 (Village)
        # Row 5: end_group
        cascade_rows = [
            r
            for r in rows
            if r[type_idx] == "select_one_from_file administration.csv"
        ]  # noqa
        self.assertEqual(len(cascade_rows), 4)

        # Level 1
        self.assertEqual(cascade_rows[0][name_idx], "location_level_1")
        self.assertEqual(cascade_rows[0][filter_idx], "level = 1")
        self.assertIsNone(cascade_rows[0][relevant_idx])
        self.assertIn("Farm Location - ", cascade_rows[0][label_idx])

        # Level 2
        self.assertEqual(cascade_rows[1][name_idx], "location_level_2")
        self.assertEqual(
            cascade_rows[1][filter_idx], "parent_key = ${location_level_1}"
        )
        self.assertEqual(
            cascade_rows[1][relevant_idx], "${location_level_1} != ''"
        )

        # Level 3
        self.assertEqual(cascade_rows[2][name_idx], "location_level_3")
        self.assertEqual(
            cascade_rows[2][filter_idx], "parent_key = ${location_level_2}"
        )
        self.assertEqual(
            cascade_rows[2][relevant_idx], "${location_level_2} != ''"
        )

    def test_attachment_allowed_file_types_body_accept(self):
        payload_attachment = {
            "id": 8001,
            "name": "Attachment Constraint Form",
            "version": 1,
            "question_group": [
                {
                    "id": 1,
                    "name": "media_group",
                    "label": "Media Group",
                    "question": [
                        {
                            "id": 1,
                            "name": "id_card_photo",
                            "label": "ID Card Photo",
                            "type": "attachment",
                            "rule": {
                                "allowedFileTypes": ["png", "jpg", "jpeg"]
                            },
                        },
                        {
                            "id": 2,
                            "name": "supporting_doc",
                            "label": "Supporting Document",
                            "type": "attachment",
                            "rule": {
                                "allowedFileTypes": ["pdf", "docx", "doc"]
                            },
                        },
                        {
                            "id": 3,
                            "name": "other_file",
                            "label": "Other File",
                            "type": "attachment",
                            "rule": {"allowedFileTypes": []},
                        },
                    ],
                }
            ],
        }
        stream, _ = generate_xlsform(payload_attachment)
        wb = openpyxl.load_workbook(stream)
        ws_survey = wb["survey"]
        headers = [cell.value for cell in ws_survey[1]]
        self.assertIn("body::accept", headers)
        accept_idx = headers.index("body::accept")
        name_idx = headers.index("name")

        rows = list(ws_survey.iter_rows(values_only=True))[1:]
        q_map = {r[name_idx]: r for r in rows if r[name_idx]}

        # Photo proof has image/*,.png,.jpg,.jpeg
        self.assertEqual(
            q_map["id_card_photo"][accept_idx], "image/*,.png,.jpg,.jpeg"
        )
        # Supporting doc has .pdf,.docx,.doc
        self.assertEqual(
            q_map["supporting_doc"][accept_idx], ".pdf,.docx,.doc"
        )
        # Unconstrained file has None
        self.assertIsNone(q_map["other_file"][accept_idx])
