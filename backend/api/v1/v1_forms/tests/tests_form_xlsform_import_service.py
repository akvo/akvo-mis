import io
import json
import tempfile
from unittest.mock import patch

import openpyxl
from django.test import TestCase

from api.v1.v1_forms.constants import FormTypes, QuestionTypes
from api.v1.v1_forms.functions import (
    normalize_form_definition,
    validate_form_definition,
)
from api.v1.v1_forms.models import Forms
from api.v1.v1_forms.services.xlsform_export import generate_xlsform
from api.v1.v1_forms.services.xlsform_import import (
    _parse_constraint,
    _parse_file_accept,
    build_form_payload,
    parse_relevant_expression,
    parse_xlsform,
    validate_preflight,
)
from api.v1.v1_forms.tasks import import_xlsform_job
from api.v1.v1_jobs.constants import JobStatus, JobTypes
from api.v1.v1_jobs.models import Jobs
from api.v1.v1_users.models import SystemUser


def _build_test_workbook(
    survey_rows,
    choices_rows=None,
    settings_row=None,
    survey_headers=None,
    choices_headers=None,
    settings_headers=None,
) -> io.BytesIO:
    """Helper to create an in-memory openpyxl workbook."""
    wb = openpyxl.Workbook()

    ws_survey = wb.active
    ws_survey.title = "survey"
    s_headers = survey_headers or [
        "type",
        "name",
        "label",
        "required",
        "hint",
        "relevant",
        "constraint",
        "appearance",
        "body::accept",
    ]
    ws_survey.append(s_headers)
    for row in survey_rows:
        ws_survey.append(row)

    if choices_rows is not None:
        ws_choices = wb.create_sheet(title="choices")
        c_headers = choices_headers or ["list_name", "name", "label"]
        ws_choices.append(c_headers)
        for row in choices_rows:
            ws_choices.append(row)

    if settings_row is not None:
        ws_settings = wb.create_sheet(title="settings")
        st_headers = settings_headers or [
            "form_title",
            "form_id",
            "version",
            "default_language",
        ]
        ws_settings.append(st_headers)
        ws_settings.append(settings_row)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


class XLSFormImportServiceTestCase(TestCase):
    def test_parse_basic_types(self):
        survey_rows = [
            ["text", "q_text", "Text", "no", None, None, None, None, None],
            ["integer", "q_int", "Int", "no", None, None, None, None, None],
            ["decimal", "q_dec", "Dec", "no", None, None, None, None, None],
            ["date", "q_date", "Date", "no", None, None, None, None, None],
            ["geopoint", "q_geo", "Geo", "no", None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        self.assertEqual(parsed["total_questions"], 5)
        self.assertEqual(len(parsed["question_groups"]), 1)
        questions = parsed["question_groups"][0]["question"]

        self.assertEqual(questions[0]["type"], "text")
        self.assertEqual(questions[0]["name"], "q_text")

        self.assertEqual(questions[1]["type"], "number")
        self.assertEqual(questions[1]["name"], "q_int")
        self.assertEqual(questions[1]["rule"], {"allowDecimal": False})

        self.assertEqual(questions[2]["type"], "number")
        self.assertEqual(questions[2]["name"], "q_dec")
        self.assertEqual(questions[2]["rule"], {"allowDecimal": True})

        self.assertEqual(questions[3]["type"], "date")
        self.assertEqual(questions[4]["type"], "geo")

    def test_parse_signature(self):
        survey_rows = [
            ["image", "q_sig", "Sig", "no", None, None, None, "signature", ""],
            ["image", "q_img", "Photo", "no", None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        questions = parsed["question_groups"][0]["question"]
        self.assertEqual(questions[0]["type"], "signature")
        self.assertEqual(questions[1]["type"], "image")

    def test_parse_select_one_and_multiple(self):
        survey_rows = [
            [
                "select_one gender_list",
                "gender",
                "Gender",
                "yes",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select_multiple hob_list",
                "hobbies",
                "Hobbies",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
        ]
        choices_rows = [
            ["gender_list", "male", "Male"],
            ["gender_list", "female", "Female"],
            ["hob_list", "reading", "Reading"],
            ["hob_list", "sports", "Sports"],
        ]
        stream = _build_test_workbook(survey_rows, choices_rows=choices_rows)
        parsed = parse_xlsform(stream)

        questions = parsed["question_groups"][0]["question"]
        self.assertEqual(questions[0]["type"], "option")
        self.assertEqual(len(questions[0]["option"]), 2)
        self.assertEqual(questions[0]["option"][0]["value"], "male")
        self.assertEqual(questions[0]["option"][0]["label"], "Male")
        self.assertEqual(questions[0]["required"], True)

        self.assertEqual(questions[1]["type"], "multiple_option")
        self.assertEqual(len(questions[1]["option"]), 2)
        self.assertEqual(questions[1]["option"][1]["value"], "sports")

    def test_parse_or_other(self):
        survey_rows = [
            [
                "select_one fruit_list or_other",
                "fruit",
                "Fruit",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
        ]
        choices_rows = [
            ["fruit_list", "apple", "Apple"],
        ]
        stream = _build_test_workbook(survey_rows, choices_rows=choices_rows)
        parsed = parse_xlsform(stream)

        questions = parsed["question_groups"][0]["question"]
        opts = questions[0]["option"]
        self.assertEqual(len(opts), 2)
        self.assertEqual(opts[0]["value"], "apple")
        self.assertEqual(opts[1]["value"], "other")
        self.assertTrue(opts[1]["other"])

    def test_parse_cascade(self):
        survey_rows = [
            [
                "select_one_from_file administration.csv",
                "admin_loc",
                "Location",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        questions = parsed["question_groups"][0]["question"]
        self.assertEqual(questions[0]["type"], "cascade")

    def test_parse_attachment_and_file_accept(self):
        survey_rows = [
            ["file", "doc", "Doc", "no", None, None, None, None, ".pdf,.docx"],
            [
                "file",
                "photo",
                "Photo",
                "no",
                None,
                None,
                None,
                None,
                "image/*,.jpg,.png",
            ],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        questions = parsed["question_groups"][0]["question"]
        self.assertEqual(questions[0]["type"], "attachment")
        self.assertEqual(
            questions[0]["rule"]["allowedFileTypes"], ["pdf", "docx"]
        )
        self.assertEqual(
            questions[1]["rule"]["allowedFileTypes"], ["jpg", "png"]
        )

    def test_parse_groups(self):
        survey_rows = [
            [
                "begin_group",
                "g_personal",
                "Personal Info",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            ["text", "name", "Full Name", "no", None, None, None, None, None],
            ["end_group", None, None, None, None, None, None, None, None],
            [
                "begin_group",
                "g_contact",
                "Contact Details",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            ["text", "email", "Email", "no", None, None, None, None, None],
            ["end_group", None, None, None, None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        self.assertEqual(len(parsed["question_groups"]), 2)
        self.assertEqual(parsed["question_groups"][0]["name"], "g_personal")
        self.assertEqual(
            parsed["question_groups"][0]["label"], "Personal Info"
        )
        self.assertEqual(len(parsed["question_groups"][0]["question"]), 1)

        self.assertEqual(parsed["question_groups"][1]["name"], "g_contact")
        self.assertEqual(len(parsed["question_groups"][1]["question"]), 1)

    def test_parse_constraint(self):
        survey_rows = [
            [
                "integer",
                "age",
                "Age",
                "no",
                None,
                None,
                ". >= 18 and . <= 65",
                None,
                None,
            ],
            [
                "decimal",
                "score",
                "Score",
                "no",
                None,
                None,
                ". >= 0.5",
                None,
                None,
            ],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        questions = parsed["question_groups"][0]["question"]
        self.assertEqual(questions[0]["rule"]["min"], 18)
        self.assertEqual(questions[0]["rule"]["max"], 65)
        self.assertEqual(questions[1]["rule"]["min"], 0.5)

    def test_parse_translations(self):
        survey_headers = [
            "type",
            "name",
            "label::English (en)",
            "label::French (fr)",
            "hint::English (en)",
            "hint::French (fr)",
        ]
        survey_rows = [
            [
                "text",
                "fav_color",
                "Favorite Color",
                "Couleur préférée",
                "Pick one",
                "Choisissez-en un",
            ],
        ]
        choices_headers = [
            "list_name",
            "name",
            "label::English (en)",
            "label::French (fr)",
        ]
        choices_rows = [
            ["color_list", "red", "Red", "Rouge"],
        ]
        settings_headers = ["form_title", "default_language"]
        settings_row = ["Multilingual Survey", "English (en)"]

        stream = _build_test_workbook(
            survey_rows,
            choices_rows=choices_rows,
            settings_row=settings_row,
            survey_headers=survey_headers,
            choices_headers=choices_headers,
            settings_headers=settings_headers,
        )
        parsed = parse_xlsform(stream)

        self.assertEqual(parsed["form_name"], "Multilingual Survey")
        self.assertEqual(parsed["default_language"], "en")
        self.assertIn("fr", parsed["languages"])

        q = parsed["question_groups"][0]["question"][0]
        self.assertEqual(q["label"], "Favorite Color")
        self.assertEqual(q["tooltip"], "Pick one")
        self.assertIsNotNone(q["translations"])
        self.assertEqual(q["translations"][0]["language"], "fr")
        self.assertEqual(q["translations"][0]["label"], "Couleur préférée")
        self.assertEqual(q["translations"][0]["tooltip"], "Choisissez-en un")

    def test_unsupported_language_warns_and_skips(self):
        """Unsupported lang codes produce warnings and are not imported."""
        survey_headers = [
            "type",
            "name",
            "label::English (en)",
            "label::Tetum (tet)",  # tet is not in locale-codes
        ]
        survey_rows = [["text", "q1", "Hello", "Olá"]]
        settings_headers = ["form_title", "default_language"]
        settings_row = ["Test Form", "English (en)"]

        stream = _build_test_workbook(
            survey_rows,
            survey_headers=survey_headers,
            settings_row=settings_row,
            settings_headers=settings_headers,
        )
        parsed = parse_xlsform(stream)

        # 'tet' language should NOT be in the imported languages list
        self.assertNotIn("tet", parsed["languages"])
        # A warning should be emitted
        warning_messages = [w["message"] for w in parsed["warnings"]]
        self.assertTrue(
            any("tet" in m or "Tetum" in m for m in warning_messages),
            f"Expected warning about 'tet' language, got: {warning_messages}",
        )

    def test_parse_relevant_selected(self):
        name_to_id = {"has_car": 1, "car_type": 2}
        expr = "selected(${has_car}, 'yes')"
        deps, rule, warn = parse_relevant_expression(expr, name_to_id)

        self.assertIsNone(warn)
        self.assertEqual(rule, "AND")
        self.assertEqual(deps, [{"id": 1, "options": ["yes"]}])

    def test_parse_relevant_selected_multi_or(self):
        name_to_id = {"has_vehicle": 1}
        expr = (
            "(selected(${has_vehicle}, 'car') or "
            "selected(${has_vehicle}, 'bike'))"
        )
        deps, rule, warn = parse_relevant_expression(expr, name_to_id)

        self.assertIsNone(warn)
        self.assertEqual(deps, [{"id": 1, "options": ["car", "bike"]}])

    def test_parse_relevant_min_max(self):
        name_to_id = {"age": 5}
        expr = "${age} >= 18 and ${age} <= 65"
        deps, rule, warn = parse_relevant_expression(expr, name_to_id)

        self.assertIsNone(warn)
        self.assertEqual(rule, "AND")
        self.assertEqual(deps, [{"id": 5, "min": 18, "max": 65}])

    def test_parse_relevant_equal_and_not_equal(self):
        name_to_id = {"status": 3, "income": 4}
        expr_eq = "${status} = 'employed'"
        deps_eq, _, warn_eq = parse_relevant_expression(expr_eq, name_to_id)
        self.assertIsNone(warn_eq)
        self.assertEqual(deps_eq, [{"id": 3, "equal": "employed"}])

        expr_neq = "${status} != 'unemployed' and string-length(${status}) > 0"
        deps_neq, _, warn_neq = parse_relevant_expression(expr_neq, name_to_id)
        self.assertIsNone(warn_neq)
        self.assertEqual(deps_neq, [{"id": 3, "notEqual": "unemployed"}])

    def test_parse_relevant_or_rule(self):
        name_to_id = {"q1": 1, "q2": 2}
        expr = "${q1} = 'yes' or ${q2} = 'yes'"
        deps, rule, warn = parse_relevant_expression(expr, name_to_id)

        self.assertIsNone(warn)
        self.assertEqual(rule, "OR")
        self.assertEqual(len(deps), 2)
        self.assertEqual(deps[0], {"id": 1, "equal": "yes"})
        self.assertEqual(deps[1], {"id": 2, "equal": "yes"})

    def test_parse_relevant_unrecognized(self):
        name_to_id = {"q1": 1}
        expr = "count-selected(${q1}) > 2"
        deps, rule, warn = parse_relevant_expression(expr, name_to_id)

        self.assertIsNotNone(warn)
        self.assertEqual(deps, [])

    def test_parse_skip_unsupported(self):
        survey_rows = [
            ["calculate", "calc_1", "Calc", None, None, None, None, None],
            ["note", "note_1", "Note", None, None, None, None, None],
            ["geoshape", "poly_1", "Poly", None, None, None, None, None],
            ["geotrace", "trace_1", "Trace", None, None, None, None, None],
            ["text", "q_real", "Real", "no", None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        self.assertEqual(parsed["total_questions"], 1)
        self.assertEqual(len(parsed["skipped_rows"]), 4)
        self.assertIn("calculate", parsed["skipped_rows"][0]["message"])
        self.assertIn("geoshape", parsed["skipped_rows"][2]["message"])
        self.assertIn("geotrace", parsed["skipped_rows"][3]["message"])

    def test_preflight_error_no_survey_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "not_survey"
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        with self.assertRaises(ValueError):
            parse_xlsform(stream)

    def test_preflight_error_no_questions(self):
        survey_rows = [
            ["calculate", "calc_1", "Calc", None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)
        errors, warnings = validate_preflight(parsed)

        self.assertTrue(
            any("No valid questions found" in e["message"] for e in errors)
        )

    def test_preflight_error_duplicate_names(self):
        survey_rows = [
            ["text", "dup_name", "Q1", "no", None, None, None, None, None],
            ["text", "dup_name", "Q2", "no", None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)
        errors, warnings = validate_preflight(parsed)

        self.assertTrue(
            any("Duplicate question name" in e["message"] for e in errors)
        )

    def test_build_form_payload(self):
        survey_rows = [
            ["text", "q1", "Q1", "no", None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        # 1. Registration
        payload_reg = build_form_payload(parsed, form_type="registration")
        self.assertEqual(payload_reg["type"], FormTypes.registration)
        self.assertIsNone(payload_reg["parent_hint"])

        # 2. Monitoring with parent_id
        payload_mon = build_form_payload(
            parsed, form_type="monitoring", parent_id=42
        )
        self.assertEqual(payload_mon["type"], FormTypes.monitoring)
        self.assertEqual(payload_mon["parent_hint"], {"id": 42})

    def test_round_trip_export_and_import(self):
        dict_payload = {
            "id": 101,
            "name": "Household Profile Form",
            "version": 2,
            "languages": ["en", "fr"],
            "defaultLanguage": "en",
            "question_group": [
                {
                    "id": 1,
                    "name": "demographics",
                    "label": "Demographics",
                    "question": [
                        {
                            "id": 10,
                            "name": "full_name",
                            "label": "Full Name",
                            "type": QuestionTypes.text,
                            "required": True,
                        },
                        {
                            "id": 20,
                            "name": "age",
                            "label": "Age",
                            "type": QuestionTypes.number,
                            "rule": {
                                "min": 18,
                                "max": 100,
                                "allowDecimal": False,
                            },
                        },
                        {
                            "id": 30,
                            "name": "gender",
                            "label": "Gender",
                            "type": QuestionTypes.option,
                            "option": [
                                {"label": "Female", "value": "female"},
                                {"label": "Male", "value": "male"},
                            ],
                        },
                    ],
                }
            ],
        }

        stream, skipped = generate_xlsform(dict_payload)
        self.assertEqual(skipped, [])

        parsed = parse_xlsform(stream)
        self.assertEqual(parsed["form_name"], "Household Profile Form")
        self.assertEqual(parsed["version"], 2)
        self.assertEqual(parsed["total_questions"], 3)

        questions = parsed["question_groups"][0]["question"]
        self.assertEqual(questions[0]["name"], "full_name")
        self.assertEqual(questions[0]["type"], "text")
        self.assertTrue(questions[0]["required"])

        self.assertEqual(questions[1]["name"], "age")
        self.assertEqual(questions[1]["type"], "number")
        self.assertEqual(questions[1]["rule"]["min"], 18)
        self.assertEqual(questions[1]["rule"]["max"], 100)

        self.assertEqual(questions[2]["name"], "gender")
        self.assertEqual(questions[2]["type"], "option")
        self.assertEqual(len(questions[2]["option"]), 2)

    def test_import_xlsform_job_success(self):
        user = SystemUser.objects.create(
            email="import_test_user@akvo.org",
            first_name="Import",
            last_name="User",
        )
        survey_rows = [
            ["text", "fav_sport", "Sport", "yes", None, None, None, None],
        ]
        stream = _build_test_workbook(
            survey_rows,
            settings_row=["Sport Survey", "sport_form", "1", "English (en)"],
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tf.write(stream.getvalue())
            tf.flush()
            temp_path = tf.name

        job = Jobs.objects.create(
            type=JobTypes.import_form,
            status=JobStatus.on_progress,
            user=user,
            info={
                "file": "sport.xlsx",
                "form_type": "registration",
                "parent_id": None,
            },
        )

        with patch("api.v1.v1_forms.tasks.download", return_value=temp_path):
            import_xlsform_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, JobStatus.done)
        res = json.loads(job.result)
        self.assertIn("form_id", res)
        self.assertEqual(res["form_name"], "Sport Survey")

        # Verify created form in DB
        created_form = Forms.objects.get(id=res["form_id"])
        self.assertEqual(created_form.name, "Sport Survey")
        self.assertEqual(created_form.type, FormTypes.registration)
        self.assertEqual(created_form.form_question_group.count(), 1)
        self.assertEqual(created_form.form_questions.count(), 1)

    def test_import_xlsform_job_read_error(self):
        user = SystemUser.objects.create(
            email="read_err_user@akvo.org",
            first_name="Read",
            last_name="User",
        )
        job = Jobs.objects.create(
            type=JobTypes.import_form,
            status=JobStatus.on_progress,
            user=user,
            info={"file": "missing.xlsx"},
        )

        with patch(
            "api.v1.v1_forms.tasks.download",
            side_effect=FileNotFoundError("File not found on storage"),
        ):
            import_xlsform_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, JobStatus.failed)
        res = json.loads(job.result)
        self.assertEqual(res[0]["code"], "file_read_error")

    def test_import_xlsform_job_preflight_error(self):
        user = SystemUser.objects.create(
            email="preflight_err_user@akvo.org",
            first_name="Preflight",
            last_name="User",
        )
        survey_rows = [
            ["calculate", "calc_only", "Calc", None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tf.write(stream.getvalue())
            tf.flush()
            temp_path = tf.name

        job = Jobs.objects.create(
            type=JobTypes.import_form,
            status=JobStatus.on_progress,
            user=user,
            info={"file": "empty.xlsx"},
        )

        with patch("api.v1.v1_forms.tasks.download", return_value=temp_path):
            import_xlsform_job(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, JobStatus.failed)
        res = json.loads(job.result)
        self.assertTrue(
            any(
                "No valid questions found" in err.get("message", "")
                for err in res
            )
        )

    def test_parse_constraint_kobo_patterns(self):
        # 1. Variable name self-references and parentheses
        r, w = _parse_constraint("${macrophytes} <= 100", "macrophytes")
        self.assertEqual(r, {"max": 100})
        self.assertIsNone(w)

        r, w = _parse_constraint("(${total_weight} <= 500)", "total_weight")
        self.assertEqual(r, {"max": 500})
        self.assertIsNone(w)

        r, w = _parse_constraint("((. <= 100))")
        self.assertEqual(r, {"max": 100})
        self.assertIsNone(w)

        # 2. Compound and clauses
        r, w = _parse_constraint(". >= 0 and . <= 100")
        self.assertEqual(r, {"min": 0, "max": 100})
        self.assertIsNone(w)

        r, w = _parse_constraint("(. <= 100 and . >= 0)")
        self.assertEqual(r, {"min": 0, "max": 100})
        self.assertIsNone(w)

        r, w = _parse_constraint("((. >= 0) and (. <= 100))")
        self.assertEqual(r, {"min": 0, "max": 100})
        self.assertIsNone(w)

        r, w = _parse_constraint("0 <= . and . <= 100")
        self.assertEqual(r, {"min": 0, "max": 100})
        self.assertIsNone(w)

        # 3. Floats and strict bounds
        r, w = _parse_constraint(". >= 0.5 and . <= 99.5")
        self.assertEqual(r, {"min": 0.5, "max": 99.5})
        self.assertIsNone(w)

        # 4. Partially unparseable regex
        r, w = _parse_constraint(". <= 100 and regex(., '^[0-9]+$')", "code_q")
        self.assertEqual(r, {"max": 100})
        self.assertIsNotNone(w)
        self.assertIn("could not be fully converted", w)

    def test_kobo_validation_criteria_import(self):
        # Test full workbook parsing with Kobo-style constraints
        survey_rows = [
            [
                "decimal",
                "macrophytes",
                "% Macrophytes",
                "no",
                None,
                None,
                "${macrophytes} <= 100",
                None,
                None,
            ],
            [
                "integer",
                "fish_weight",
                "Total weight of fish caught",
                "no",
                None,
                None,
                "(${fish_weight} <= 500)",
                None,
                None,
            ],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)

        q1 = parsed["question_groups"][0]["question"][0]
        self.assertEqual(q1["name"], "macrophytes")
        self.assertEqual(q1["rule"]["max"], 100)
        self.assertTrue(q1["rule"]["allowDecimal"])

        q2 = parsed["question_groups"][0]["question"][1]
        self.assertEqual(q2["name"], "fish_weight")
        self.assertEqual(q2["rule"]["max"], 500)
        self.assertFalse(q2["rule"]["allowDecimal"])

    def test_repeat_group_dynamic_repeat_count_warning(self):
        s_headers = [
            "type",
            "name",
            "label",
            "required",
            "hint",
            "relevant",
            "constraint",
            "appearance",
            "body::accept",
            "repeat_count",
        ]
        survey_rows = [
            [
                "select_multiple species",
                "species_list",
                "Observed Species",
                "no",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "begin_repeat",
                "species_repeat",
                "Species Details",
                None,
                None,
                None,
                None,
                None,
                None,
                "count-selected(${species_list})",
            ],
            [
                "decimal",
                "percentage_value",
                "% value",
                "no",
                None,
                None,
                ". <= 100",
                None,
                None,
                None,
            ],
            [
                "end_repeat",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ]
        choices_rows = [
            ["species", "tilapia", "Tilapia"],
            ["species", "catfish", "Catfish"],
        ]
        stream = _build_test_workbook(
            survey_rows,
            choices_rows=choices_rows,
            survey_headers=s_headers,
        )
        parsed = parse_xlsform(stream)
        self.assertTrue(
            any(
                "dynamic repeat_count" in w.get("message", "")
                for w in parsed["warnings"]
            )
        )
        groups = parsed["question_groups"]
        repeat_group = [g for g in groups if g["name"] == "species_repeat"][0]
        self.assertTrue(repeat_group["repeatable"])

    def test_group_level_relevant_warning(self):
        survey_rows = [
            [
                "select_one yn",
                "has_children",
                "Have children?",
                "yes",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "begin_group",
                "child_group",
                "Child Details",
                None,
                None,
                "selected(${has_children}, 'yes')",
                None,
                None,
                None,
            ],
            [
                "integer",
                "num_children",
                "Number of Children",
                "yes",
                None,
                None,
                ". >= 1",
                None,
                None,
            ],
            ["end_group", None, None, None, None, None, None, None, None],
        ]
        choices_rows = [
            ["yn", "yes", "Yes"],
            ["yn", "no", "No"],
        ]
        stream = _build_test_workbook(
            survey_rows,
            choices_rows=choices_rows,
        )
        parsed = parse_xlsform(stream)
        self.assertTrue(
            any(
                "Group 'child_group' has a 'relevant' condition"
                in w.get("message", "")
                for w in parsed["warnings"]
            )
        )

    def test_calculate_type_and_calculation_column_warnings(self):
        s_headers = [
            "type",
            "name",
            "label",
            "required",
            "hint",
            "relevant",
            "constraint",
            "appearance",
            "body::accept",
            "calculation",
        ]
        survey_rows = [
            [
                "integer",
                "fish_a",
                "Fish A count",
                "no",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "calculate",
                "total_calc",
                "Total",
                None,
                None,
                None,
                None,
                None,
                None,
                "${fish_a} * 2",
            ],
            [
                "integer",
                "fish_b",
                "Fish B count",
                "no",
                None,
                None,
                None,
                None,
                None,
                "${fish_a} + 5",
            ],
        ]
        stream = _build_test_workbook(
            survey_rows,
            survey_headers=s_headers,
        )
        parsed = parse_xlsform(stream)
        self.assertEqual(parsed["total_questions"], 2)  # calculate skipped
        self.assertTrue(
            any(
                "Calculated field 'total_calc'" in w.get("message", "")
                for w in parsed["warnings"]
            )
        )
        self.assertTrue(
            any(
                "Question 'fish_b' has calculation" in w.get("message", "")
                for w in parsed["warnings"]
            )
        )

    def test_parse_all_question_types_and_aliases(self):
        """Test parsing of all supported question types and their aliases."""
        survey_rows = [
            ["text", "q_text", "Text", "no", None, None, None, None, None],
            ["string", "q_str", "Str", "no", None, None, None, None, None],
            ["input", "q_inp", "Inp", "no", None, None, None, None, None],
            ["integer", "q_int", "Int", "no", None, None, None, None, None],
            ["int", "q_int_alias", "IntA", "no", None, None, None, None, None],
            ["decimal", "q_dec", "Dec", "no", None, None, None, None, None],
            ["date", "q_date", "Date", "no", None, None, None, None, None],
            ["geopoint", "q_geo", "Geo", "no", None, None, None, None, None],
            ["image", "q_img", "Img", "no", None, None, None, None, None],
            ["photo", "q_photo", "Photo", "no", None, None, None, None, None],
            [
                "image",
                "q_sig",
                "Sig",
                "no",
                None,
                None,
                None,
                "signature",
                None,
            ],
            [
                "photo",
                "q_sig2",
                "Sig2",
                "no",
                None,
                None,
                None,
                "signature",
                None,
            ],
            [
                "file",
                "q_att",
                "Att",
                "no",
                None,
                None,
                None,
                None,
                ".pdf,.docx",
            ],
            [
                "select_one_from_file administration.csv",
                "q_casc",
                "Casc",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select one_from_file administration.csv",
                "q_casc2",
                "Casc2",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select_one fruits",
                "q_opt",
                "Opt",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select one fruits",
                "q_opt2",
                "Opt2",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select_1 fruits",
                "q_opt3",
                "Opt3",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select_one fruits or_other",
                "q_opt_other",
                "OptO",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select_multiple colors",
                "q_mopt",
                "MOpt",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select multiple colors",
                "q_mopt2",
                "MOpt2",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "select_multiple colors or_other",
                "q_mopt_other",
                "MOptO",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
        ]
        choices_rows = [
            ["fruits", "apple", "Apple"],
            ["fruits", "banana", "Banana"],
            ["colors", "red", "Red"],
            ["colors", "blue", "Blue"],
        ]
        stream = _build_test_workbook(survey_rows, choices_rows=choices_rows)
        parsed = parse_xlsform(stream)

        self.assertEqual(parsed["total_questions"], len(survey_rows))
        q_map = {
            q["name"]: q for q in parsed["question_groups"][0]["question"]
        }

        self.assertEqual(q_map["q_text"]["type"], "text")
        self.assertEqual(q_map["q_str"]["type"], "text")
        self.assertEqual(q_map["q_inp"]["type"], "text")

        self.assertEqual(q_map["q_int"]["type"], "number")
        self.assertEqual(q_map["q_int"]["rule"]["allowDecimal"], False)
        self.assertEqual(q_map["q_int_alias"]["type"], "number")
        self.assertEqual(q_map["q_int_alias"]["rule"]["allowDecimal"], False)

        self.assertEqual(q_map["q_dec"]["type"], "number")
        self.assertEqual(q_map["q_dec"]["rule"]["allowDecimal"], True)

        self.assertEqual(q_map["q_date"]["type"], "date")
        self.assertEqual(q_map["q_geo"]["type"], "geo")

        self.assertEqual(q_map["q_img"]["type"], "image")
        self.assertEqual(q_map["q_photo"]["type"], "image")
        self.assertEqual(q_map["q_sig"]["type"], "signature")
        self.assertEqual(q_map["q_sig2"]["type"], "signature")

        self.assertEqual(q_map["q_att"]["type"], "attachment")
        self.assertEqual(
            q_map["q_att"]["rule"]["allowedFileTypes"], ["pdf", "docx"]
        )

        self.assertEqual(q_map["q_casc"]["type"], "cascade")
        self.assertEqual(q_map["q_casc2"]["type"], "cascade")

        self.assertEqual(q_map["q_opt"]["type"], "option")
        self.assertEqual(len(q_map["q_opt"]["option"]), 2)
        self.assertEqual(q_map["q_opt2"]["type"], "option")
        self.assertEqual(q_map["q_opt3"]["type"], "option")

        self.assertEqual(q_map["q_opt_other"]["type"], "option")
        self.assertEqual(len(q_map["q_opt_other"]["option"]), 3)
        self.assertTrue(q_map["q_opt_other"]["option"][-1]["other"])

        self.assertEqual(q_map["q_mopt"]["type"], "multiple_option")
        self.assertEqual(len(q_map["q_mopt"]["option"]), 2)
        self.assertEqual(q_map["q_mopt2"]["type"], "multiple_option")

        self.assertEqual(q_map["q_mopt_other"]["type"], "multiple_option")
        self.assertEqual(len(q_map["q_mopt_other"]["option"]), 3)
        self.assertTrue(q_map["q_mopt_other"]["option"][-1]["other"])

    def test_parse_constraint_exhaustive_side_conditions(self):
        """Test boundary, chained, strict inequality, decimal constraints."""
        # 1. Chained comparisons
        rule, warn = _parse_constraint("0 <= . <= 100", "q")
        self.assertEqual(rule, {"min": 0, "max": 100})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("100 >= . >= 0", "q")
        self.assertEqual(rule, {"min": 0, "max": 100})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint(
            "0 <= ${macrophytes} <= 100", "macrophytes"
        )
        self.assertEqual(rule, {"min": 0, "max": 100})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("0 < . < 500", "q")
        self.assertEqual(rule, {"min": 0, "max": 500})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("500 > . > 0", "q")
        self.assertEqual(rule, {"min": 0, "max": 500})
        self.assertIsNone(warn)

        # 2. Strict single inequalities
        rule, warn = _parse_constraint(". > 0", "q")
        self.assertEqual(rule, {"min": 0})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint(". < 500", "q")
        self.assertEqual(rule, {"max": 500})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("0 < .", "q")
        self.assertEqual(rule, {"min": 0})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("500 > .", "q")
        self.assertEqual(rule, {"max": 500})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("${weight} > 0", "weight")
        self.assertEqual(rule, {"min": 0})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("500 > ${weight}", "weight")
        self.assertEqual(rule, {"max": 500})
        self.assertIsNone(warn)

        # 3. Decimals and Negatives
        rule, warn = _parse_constraint("-50.5 <= . and . <= 50.5", "q")
        self.assertEqual(rule, {"min": -50.5, "max": 50.5})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("(. >= -10.25) and (. <= 99.75)", "q")
        self.assertEqual(rule, {"min": -10.25, "max": 99.75})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("${temp} >= -273.15", "temp")
        self.assertEqual(rule, {"min": -273.15})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("${temp} <= 1000.5", "temp")
        self.assertEqual(rule, {"max": 1000.5})
        self.assertIsNone(warn)

        # 4. Spaces and nested parentheses
        rule, warn = _parse_constraint(
            "( ( . >= 0 ) ) and ( ( . <= 100 ) )", "q"
        )
        self.assertEqual(rule, {"min": 0, "max": 100})
        self.assertIsNone(warn)

        rule, warn = _parse_constraint("  .  >=  10  and  .  <=  90  ", "q")
        self.assertEqual(rule, {"min": 10, "max": 90})
        self.assertIsNone(warn)

        # 5. Case-insensitive AND
        rule, warn = _parse_constraint(". >= 10 AND . <= 20", "q")
        self.assertEqual(rule, {"min": 10, "max": 20})
        self.assertIsNone(warn)

        # 6. Compound with partial non-numeric clause
        rule, warn = _parse_constraint(". >= 0 and . <= 100 and . != 50", "q")
        self.assertEqual(rule, {"min": 0, "max": 100})
        self.assertIsNotNone(warn)
        self.assertIn(". != 50", warn)

        # 7. Unparseable constraints
        rule, warn = _parse_constraint("regex(., '^[A-Z]{3}$')", "q")
        self.assertEqual(rule, {})
        self.assertIsNotNone(warn)
        self.assertIn("could not be", warn)

        rule, warn = _parse_constraint("string-length(.) <= 250", "q")
        self.assertEqual(rule, {})
        self.assertIsNotNone(warn)

        rule, warn = _parse_constraint("count-selected(.) >= 1", "q")
        self.assertEqual(rule, {})
        self.assertIsNotNone(warn)
        self.assertIn("could not be", warn)

        # 8. Empty / Whitespace / None
        self.assertEqual(_parse_constraint("", "q"), ({}, None))
        self.assertEqual(_parse_constraint("   ", "q"), ({}, None))
        self.assertEqual(_parse_constraint(None, "q"), ({}, None))

    def test_parse_file_accept_side_conditions(self):
        """Test parsing of MIME/extension strings from body::accept."""
        self.assertEqual(
            _parse_file_accept("image/*, .pdf, .docx"), ["pdf", "docx"]
        )
        self.assertEqual(
            _parse_file_accept("IMAGE/*, .PDF, .DOCX"), ["pdf", "docx"]
        )
        self.assertEqual(
            _parse_file_accept("pdf, docx, xlsx"), ["pdf", "docx", "xlsx"]
        )
        self.assertEqual(
            _parse_file_accept(" .pdf , , .jpg  , image/* "), ["pdf", "jpg"]
        )
        self.assertEqual(_parse_file_accept(""), [])
        self.assertEqual(_parse_file_accept("   "), [])
        self.assertEqual(_parse_file_accept(None), [])

    def test_required_flag_variations(self):
        """Test truthy and falsy variations of required column."""
        survey_rows = [
            ["text", "q_yes", "Q1", "yes", None, None, None, None, None],
            ["text", "q_true", "Q2", "true", None, None, None, None, None],
            ["text", "q_one", "Q3", "1", None, None, None, None, None],
            ["text", "q_True", "Q4", "True", None, None, None, None, None],
            ["text", "q_YES", "Q5", "YES", None, None, None, None, None],
            ["text", "q_no", "Q6", "no", None, None, None, None, None],
            ["text", "q_false", "Q7", "false", None, None, None, None, None],
            ["text", "q_zero", "Q8", "0", None, None, None, None, None],
            ["text", "q_empty", "Q9", "", None, None, None, None, None],
            ["text", "q_none", "Q10", None, None, None, None, None, None],
        ]
        stream = _build_test_workbook(survey_rows)
        parsed = parse_xlsform(stream)
        questions = parsed["question_groups"][0]["question"]

        for q in questions[:5]:
            self.assertTrue(q["required"], f"Failed for {q['name']}")
            self.assertTrue(
                (q.get("rule") or {}).get("required"),
                f"Rule required failed for {q['name']}",
            )

        for q in questions[5:]:
            self.assertFalse(q["required"], f"Failed for {q['name']}")
            self.assertFalse(
                (q.get("rule") or {}).get("required", False),
                f"Rule required should be False for {q['name']}",
            )

    def test_repeat_group_leading_question_roundtrip(self):
        """Test repeat group linking to leading_question via repeat_count."""
        survey_rows = [
            [
                "select_multiple species",
                "fish_species",
                "Observed Species",
                "no",
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "begin_repeat",
                "species_repeat",
                "Species Details",
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            [
                "decimal",
                "percent_value",
                "% Value",
                "no",
                None,
                None,
                "0 <= . <= 100",
                None,
                None,
            ],
            [
                "end_repeat",
                "species_repeat",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ]
        survey_headers = [
            "type",
            "name",
            "label",
            "required",
            "hint",
            "relevant",
            "constraint",
            "appearance",
            "body::accept",
            "repeat_count",
        ]
        # Append repeat_count to begin_repeat row
        survey_rows[0].append(None)
        survey_rows[1].append("count-selected(${fish_species})")
        survey_rows[2].append(None)
        survey_rows[3].append(None)

        choices_rows = [
            ["species", "tilapia", "Tilapia"],
            ["species", "catfish", "Catfish"],
            ["species", "perch", "Perch"],
        ]
        stream = _build_test_workbook(
            survey_rows,
            choices_rows=choices_rows,
            survey_headers=survey_headers,
        )
        parsed = parse_xlsform(stream)
        self.assertEqual(parsed["total_questions"], 2)
        self.assertEqual(len(parsed["question_groups"]), 2)

        repeat_group = parsed["question_groups"][1]
        self.assertTrue(repeat_group["repeatable"])
        self.assertEqual(repeat_group["repeat_text"], "Add another")
        # Leading question tmp_id should point to fish_species (tmp_id 1)
        self.assertEqual(repeat_group["leading_question"], 1)

        # Preflight validation passes with non-blocking warning
        errors, warnings = validate_preflight(parsed)
        self.assertEqual(errors, [])
        self.assertTrue(
            any(
                "Repeat group 'species_repeat'" in w["message"]
                for w in warnings
            )
        )

        # Normalize and validate
        payload = build_form_payload(parsed)
        norm = normalize_form_definition(payload)
        issues = validate_form_definition(norm, check_entities=False)
        blocking_errors = [i for i in issues if i.get("level") == "error"]
        self.assertEqual(blocking_errors, [])

    def test_relevant_expression_exhaustive_side_conditions(self):
        """Test skip-logic parsing for quotes, chained, equality, etc."""
        name_map = {
            "gender": 1,
            "age": 2,
            "status": 3,
            "score": 4,
            "temp": 5,
        }

        # 1. Single vs double quotes in selected()
        deps, rule, warn = parse_relevant_expression(
            "selected(${gender}, 'female')", name_map
        )
        self.assertEqual(deps, [{"id": 1, "options": ["female"]}])
        self.assertEqual(rule, "AND")
        self.assertIsNone(warn)

        deps, rule, warn = parse_relevant_expression(
            'selected(${gender}, "female")', name_map
        )
        self.assertEqual(deps, [{"id": 1, "options": ["female"]}])
        self.assertEqual(rule, "AND")
        self.assertIsNone(warn)

        deps, rule, warn = parse_relevant_expression(
            'selected(  ${gender}  ,  "female"  )', name_map
        )
        self.assertEqual(deps, [{"id": 1, "options": ["female"]}])
        self.assertEqual(rule, "AND")
        self.assertIsNone(warn)

        # 2. Combined selected options with OR
        deps, rule, warn = parse_relevant_expression(
            "(selected(${gender}, 'female') or "
            "selected(${gender}, \"other\"))",
            name_map,
        )
        self.assertEqual(deps, [{"id": 1, "options": ["female", "other"]}])
        self.assertEqual(rule, "OR")
        self.assertIsNone(warn)

        # 3. Numeric min and max
        deps, rule, warn = parse_relevant_expression(
            "${age} >= 18 and ${age} <= 65", name_map
        )
        self.assertEqual(deps, [{"id": 2, "min": 18, "max": 65}])
        self.assertEqual(rule, "AND")
        self.assertIsNone(warn)

        deps, rule, warn = parse_relevant_expression(
            "${temp} >= -10.5 and ${temp} <= 45.5", name_map
        )
        self.assertEqual(deps, [{"id": 5, "min": -10.5, "max": 45.5}])
        self.assertIsNone(warn)

        # 4. Equality & Inequality with single and double quotes
        deps, rule, warn = parse_relevant_expression(
            "${status} = 'active'", name_map
        )
        self.assertEqual(deps, [{"id": 3, "equal": "active"}])
        self.assertIsNone(warn)

        deps, rule, warn = parse_relevant_expression(
            '${status} = "active"', name_map
        )
        self.assertEqual(deps, [{"id": 3, "equal": "active"}])
        self.assertIsNone(warn)

        deps, rule, warn = parse_relevant_expression(
            "${status} != 'inactive' and string-length(${status}) > 0",
            name_map,
        )
        self.assertEqual(deps, [{"id": 3, "notEqual": "inactive"}])
        self.assertIsNone(warn)

        deps, rule, warn = parse_relevant_expression(
            '${status} != "inactive"', name_map
        )
        self.assertEqual(deps, [{"id": 3, "notEqual": "inactive"}])
        self.assertIsNone(warn)

        # 5. Unknown question references
        deps, rule, warn = parse_relevant_expression(
            "selected(${missing_q}, 'val')", name_map
        )
        self.assertEqual(deps, [])
        self.assertIsNotNone(warn)
        self.assertIn("missing_q", warn)

        # 6. Unparseable XPath functions
        deps, rule, warn = parse_relevant_expression(
            "today() > date('2020-01-01')", name_map
        )
        self.assertEqual(deps, [])
        self.assertIsNotNone(warn)

    def test_full_export_import_roundtrip_all_question_types(self):
        """End-to-end test verifying all question types bidirectionally."""
        dict_payload = {
            "name": "Comprehensive Parity Test Form",
            "type": FormTypes.registration,
            "version": 1,
            "languages": ["en"],
            "default_language": "en",
            "question_group": [
                {
                    "name": "general_info",
                    "label": "General Info",
                    "order": 1,
                    "repeatable": False,
                    "question": [
                        {
                            "id": 1,
                            "name": "respondent_name",
                            "label": "Full Name",
                            "type": QuestionTypes.FieldStr[QuestionTypes.text],
                            "required": True,
                            "tooltip": {"text": "Enter legal name"},
                        },
                        {
                            "id": 2,
                            "name": "respondent_age",
                            "label": "Age",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.number
                            ],
                            "required": True,
                            "rule": {
                                "min": 18,
                                "max": 120,
                                "allowDecimal": False,
                            },
                        },
                        {
                            "id": 3,
                            "name": "body_temperature",
                            "label": "Temperature (°C)",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.number
                            ],
                            "required": False,
                            "rule": {
                                "min": 35.0,
                                "max": 43.0,
                                "allowDecimal": True,
                            },
                        },
                        {
                            "id": 4,
                            "name": "interview_date",
                            "label": "Date of Interview",
                            "type": QuestionTypes.FieldStr[QuestionTypes.date],
                            "required": True,
                        },
                        {
                            "id": 5,
                            "name": "location_point",
                            "label": "Location GPS",
                            "type": QuestionTypes.FieldStr[QuestionTypes.geo],
                            "required": False,
                        },
                        {
                            "id": 6,
                            "name": "household_type",
                            "label": "Household Type",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.option
                            ],
                            "required": True,
                            "option": [
                                {"label": "Urban", "value": "urban"},
                                {"label": "Rural", "value": "rural"},
                            ],
                        },
                        {
                            "id": 7,
                            "name": "observed_assets",
                            "label": "Observed Assets",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.multiple_option
                            ],
                            "required": False,
                            "option": [
                                {"label": "Bicycle", "value": "bike"},
                                {"label": "Solar Panel", "value": "solar"},
                            ],
                        },
                        {
                            "id": 8,
                            "name": "admin_region",
                            "label": "Administrative Region",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.cascade
                            ],
                            "required": False,
                        },
                        {
                            "id": 9,
                            "name": "national_id_doc",
                            "label": "National ID Document",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.attachment
                            ],
                            "required": False,
                            "rule": {
                                "allowedFileTypes": ["pdf", "jpg", "png"]
                            },
                        },
                        {
                            "id": 10,
                            "name": "house_photo",
                            "label": "Photo of House",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.image
                            ],
                            "required": False,
                        },
                        {
                            "id": 11,
                            "name": "respondent_signature",
                            "label": "Signature",
                            "type": QuestionTypes.FieldStr[
                                QuestionTypes.signature
                            ],
                            "required": True,
                        },
                    ],
                }
            ],
        }

        # 1. Export to XLSForm workbook
        stream, skipped = generate_xlsform(dict_payload)
        self.assertEqual(skipped, [])

        # 2. Parse workbook with parse_xlsform
        # (cascade expands to 4 admin levels: 10 + 4 = 14)
        parsed = parse_xlsform(stream)
        self.assertEqual(
            parsed["form_name"], "Comprehensive Parity Test Form"
        )
        self.assertEqual(parsed["total_questions"], 14)

        # 3. Preflight check
        errors, warnings = validate_preflight(parsed)
        self.assertEqual(errors, [])

        # 4. Build canonical definition
        payload = build_form_payload(parsed)
        self.assertEqual(payload["name"], "Comprehensive Parity Test Form")

        # 5. Normalize & validate using standard form definition validator
        norm = normalize_form_definition(payload)
        issues = validate_form_definition(norm, check_entities=False)
        blocking_errors = [i for i in issues if i.get("level") == "error"]
        self.assertEqual(
            blocking_errors,
            [],
            f"Blocking validation issues found: {blocking_errors}",
        )

        # 6. Verify question types and attributes on imported questions
        imported_qs = {
            q["name"]: q for q in norm["question_group"][0]["question"]
        }
        self.assertEqual(imported_qs["respondent_name"]["type"], "text")
        self.assertTrue(imported_qs["respondent_name"]["required"])

        self.assertEqual(imported_qs["respondent_age"]["type"], "number")
        self.assertEqual(
            imported_qs["respondent_age"]["rule"]["allowDecimal"], False
        )
        self.assertEqual(imported_qs["respondent_age"]["rule"]["min"], 18)
        self.assertEqual(imported_qs["respondent_age"]["rule"]["max"], 120)

        self.assertEqual(imported_qs["body_temperature"]["type"], "number")
        self.assertEqual(
            imported_qs["body_temperature"]["rule"]["allowDecimal"], True
        )
        self.assertEqual(imported_qs["body_temperature"]["rule"]["min"], 35.0)
        self.assertEqual(imported_qs["body_temperature"]["rule"]["max"], 43.0)

        self.assertEqual(imported_qs["interview_date"]["type"], "date")
        self.assertEqual(imported_qs["location_point"]["type"], "geo")
        self.assertEqual(imported_qs["household_type"]["type"], "option")
        self.assertEqual(
            imported_qs["observed_assets"]["type"], "multiple_option"
        )
        cascade_qs = [
            q
            for q in norm["question_group"][0]["question"]
            if q["type"] == "cascade"
        ]
        self.assertTrue(len(cascade_qs) >= 1)
        for cq in cascade_qs:
            self.assertEqual(cq["type"], "cascade")
            self.assertTrue(cq["name"].startswith("admin_region"))

        self.assertEqual(imported_qs["national_id_doc"]["type"], "attachment")
        self.assertEqual(
            imported_qs["national_id_doc"]["rule"]["allowedFileTypes"],
            ["pdf", "jpg", "png"],
        )
        self.assertEqual(imported_qs["house_photo"]["type"], "image")
        self.assertEqual(
            imported_qs["respondent_signature"]["type"], "signature"
        )

    def test_invalid_language_tag_emits_warning_and_is_excluded(self):
        """Language tags not in locale-codes (e.g. 'ns') must produce a
        preflight warning and must NOT appear in the languages list or any
        translation — preventing a crash in akvo-react-form's transformForm
        (getByTag(x) would return undefined and .name would throw TypeError).
        """
        stream = _build_test_workbook(
            survey_headers=[
                "type",
                "name",
                "label::English (en)",
                "label::Nso (ns)",
            ],
            survey_rows=[
                ["text", "q_name", "Full Name", "Jina Elfu"],
            ],
            settings_row=["My Form", "my_form", "1", "English (en)"],
        )
        parsed = parse_xlsform(stream)

        # 'ns' is invalid — should NOT appear in languages
        self.assertNotIn("ns", parsed["languages"])
        # 'en' (default) should still be present
        self.assertIn("en", parsed["languages"])

        # A warning must be emitted naming the bad tag
        warns = [w["message"] for w in parsed.get("warnings", [])]
        bad_tag_warns = [w for w in warns if "'ns'" in w]
        self.assertTrue(
            bad_tag_warns,
            "Expected a warning about unsupported language tag 'ns'",
        )

        # Translations for 'ns' must not be written to questions
        q = parsed["question_groups"][0]["question"][0]
        for trans in q.get("translations") or []:
            self.assertNotEqual(
                trans.get("language"), "ns",
                "Translation for invalid lang 'ns' must not be imported",
            )
