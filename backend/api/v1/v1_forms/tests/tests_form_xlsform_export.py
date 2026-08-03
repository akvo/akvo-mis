import openpyxl
from unittest.mock import MagicMock

from django.core.management import call_command
from django.db import connection
from django.test import TestCase
from django.test.utils import override_settings
from api.v1.v1_forms.constants import QuestionTypes
from api.v1.v1_forms.models import Forms
from api.v1.v1_forms.services.xlsform_export import (
    _map_type,
    _build_question_map,
    _build_settings_row,
    _build_choices_rows,
    _build_survey_rows,
    _build_relevant_expression,
    _build_constraint,
    generate_xlsform,
)


class DummyObject:
    def __init__(self, **kwargs):
        self.rule = None
        self.tooltip = None
        self.translations = None
        self.api = None
        for k, v in kwargs.items():
            setattr(self, k, v)


class XLSFormExportServiceTestCase(TestCase):
    def test_map_type_all_17_types(self):
        # 1. Input / Text
        self.assertEqual(
            _map_type(
                DummyObject(type=QuestionTypes.input, rule=None, name="t1")
            ),
            ("text", None),
        )
        self.assertEqual(
            _map_type(
                DummyObject(type=QuestionTypes.text, rule=None, name="t2")
            ),
            ("text", None),
        )

        # 2. Number (Integer vs Decimal)
        self.assertEqual(
            _map_type(
                DummyObject(
                    type=QuestionTypes.number,
                    rule={"allowDecimal": False},
                    name="n1",
                )
            ),
            ("integer", None),
        )
        self.assertEqual(
            _map_type(
                DummyObject(
                    type=QuestionTypes.number,
                    rule={"allowDecimal": True},
                    name="n2",
                )
            ),
            ("decimal", None),
        )
        self.assertEqual(
            _map_type(
                DummyObject(type=QuestionTypes.number, rule=None, name="n3")
            ),
            ("integer", None),
        )

        # 3. Date
        self.assertEqual(
            _map_type(
                DummyObject(type=QuestionTypes.date, rule=None, name="d1")
            ),
            ("date", None),
        )

        # 4. Option (Single select) without and with 'other'
        opt_normal = DummyObject(other=False)
        opt_other = DummyObject(other=True)
        q_opt1 = DummyObject(
            type=QuestionTypes.option,
            name="q_opt1",
            question_question_option=DummyObject(all=lambda: [opt_normal]),
        )
        q_opt2 = DummyObject(
            type=QuestionTypes.option,
            name="q_opt2",
            question_question_option=DummyObject(
                all=lambda: [opt_normal, opt_other]
            ),
        )
        self.assertEqual(_map_type(q_opt1), ("select_one option_q_opt1", None))
        self.assertEqual(
            _map_type(q_opt2), ("select_one option_q_opt2 or_other", None)
        )

        # 5. Multiple Option without and with 'other'
        q_mopt1 = DummyObject(
            type=QuestionTypes.multiple_option,
            name="q_mopt1",
            question_question_option=DummyObject(all=lambda: [opt_normal]),
        )
        q_mopt2 = DummyObject(
            type=QuestionTypes.multiple_option,
            name="q_mopt2",
            question_question_option=DummyObject(
                all=lambda: [opt_normal, opt_other]
            ),
        )
        self.assertEqual(
            _map_type(q_mopt1), ("select_multiple option_q_mopt1", None)
        )
        self.assertEqual(
            _map_type(q_mopt2),
            ("select_multiple option_q_mopt2 or_other", None),
        )

        # 6. Geo & Geo shapes
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.geo, name="g1")),
            ("geopoint", None),
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.geoshape, name="g2")),
            ("geoshape", None),
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.geotrace, name="g3")),
            ("geotrace", None),
        )

        # 7. Image & Attachment
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.image, name="i1")),
            ("image", None),
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.attachment, name="a1")),
            ("file", None),
        )

        # 8. Signature
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.signature, name="s1")),
            ("image", "signature"),
        )

        # 9. Cascade
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.cascade, name="c1")),
            ("select_one_from_file administration.csv", None),
        )

        # 10. Skipped Types (Tree, Table, Autofield)
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.tree, name="tr")),
            (None, None),
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.table, name="tb")),
            (None, None),
        )
        self.assertEqual(
            _map_type(DummyObject(type=QuestionTypes.autofield, name="af")),
            (None, None),
        )

    def test_build_question_map(self):
        q1 = DummyObject(id=10, name="q_one", type=QuestionTypes.input)
        q2 = DummyObject(id=20, name="q_two", type=QuestionTypes.option)
        group = DummyObject(question_group_question=MagicMock())
        group.question_group_question.all.return_value = [q1, q2]

        form = DummyObject(form_question_group=MagicMock())
        form.form_question_group.all.return_value = [group]

        qmap = _build_question_map(form)
        self.assertEqual(
            qmap,
            {
                10: {"name": "q_one", "type": QuestionTypes.input},
                20: {"name": "q_two", "type": QuestionTypes.option},
            },
        )

    def test_build_settings_row(self):
        form = DummyObject(
            id=123,
            name="Test Form",
            version=2,
            default_language="en",
        )
        settings = _build_settings_row(form)
        self.assertEqual(
            settings,
            {
                "form_title": "Test Form",
                "form_id": "form_123",
                "version": "2",
                "default_language": "en",
            },
        )

    def test_build_choices_rows_with_translations_and_other_filter(self):
        opt1 = DummyObject(
            id=1,
            value="yes",
            label="Yes",
            other=False,
            translations={"fr": {"label": "Oui"}},
        )
        opt2 = DummyObject(
            id=2,
            value="no",
            label="No",
            other=False,
            translations={"fr": {"label": "Non"}},
        )
        opt_other = DummyObject(
            id=3, value="other", label="Other", other=True, translations=None
        )

        q_opt = DummyObject(
            id=10,
            name="decision",
            type=QuestionTypes.option,
            question_question_option=DummyObject(
                all=lambda: [opt1, opt2, opt_other]
            ),
        )
        group = DummyObject(
            question_group_question=MagicMock(all=lambda: [q_opt])
        )
        form = DummyObject(form_question_group=MagicMock(all=lambda: [group]))

        choices = _build_choices_rows(form, lang_cols=["fr"])
        self.assertEqual(len(choices), 2)
        self.assertEqual(
            choices[0],
            {
                "list_name": "option_decision",
                "name": "yes",
                "label": "Yes",
                "label::fr": "Oui",
            },
        )
        self.assertEqual(
            choices[1],
            {
                "list_name": "option_decision",
                "name": "no",
                "label": "No",
                "label::fr": "Non",
            },
        )

    def test_build_survey_rows_repeatable_group_and_multilingual_hints(self):
        q1 = DummyObject(
            id=101,
            name="member_name",
            type=QuestionTypes.input,
            label="Member Name",
            required=True,
            tooltip={"text": "Enter full name"},
            translations={
                "fr": {
                    "label": "Nom du membre",
                    "tooltip": {"text": "Entrez le nom complet"},
                }
            },
            rule=None,
            api=None,
        )
        q_skipped = DummyObject(
            id=102,
            name="tree_select",
            type=QuestionTypes.tree,
            label="Tree",
            required=False,
            tooltip=None,
            translations=None,
            rule=None,
            api=None,
        )

        group_repeat = DummyObject(
            id=1,
            name="household_members",
            label="Household Members",
            repeatable=True,
            translations={"fr": {"label": "Membres du ménage"}},
            question_group_question=MagicMock(all=lambda: [q1, q_skipped]),
        )

        form = DummyObject(
            form_question_group=MagicMock(all=lambda: [group_repeat])
        )

        survey_rows, skipped = _build_survey_rows(
            form, question_map={}, lang_cols=["fr"]
        )

        self.assertEqual(skipped, ["tree_select"])
        self.assertEqual(len(survey_rows), 3)

        # 1. begin_repeat
        self.assertEqual(
            survey_rows[0],
            {
                "type": "begin_repeat",
                "name": "household_members",
                "label": "Household Members",
                "label::fr": "Membres du ménage",
            },
        )

        # 2. question row
        self.assertEqual(
            survey_rows[1],
            {
                "type": "text",
                "name": "member_name",
                "label": "Member Name",
                "required": "yes",
                "hint": "Enter full name",
                "label::fr": "Nom du membre",
                "hint::fr": "Entrez le nom complet",
            },
        )

        # 3. end_repeat
        self.assertEqual(survey_rows[2], {"type": "end_repeat"})

    def test_generate_xlsform_full_excel_verification(self):
        opt1 = DummyObject(
            id=1, value="m", label="Male", other=False, translations=None
        )
        opt2 = DummyObject(
            id=2, value="f", label="Female", other=False, translations=None
        )
        q_gender = DummyObject(
            id=1,
            name="gender",
            type=QuestionTypes.option,
            label="Gender",
            required=True,
            tooltip=None,
            translations=None,
            rule=None,
            api=None,
            question_question_option=DummyObject(all=lambda: [opt1, opt2]),
        )
        group = DummyObject(
            id=10,
            name="demographics",
            label="Demographics",
            repeatable=False,
            translations=None,
            question_group_question=MagicMock(all=lambda: [q_gender]),
        )
        form = DummyObject(
            id=99,
            name="Census Survey",
            version=3,
            default_language="en",
            languages=["en", "fr"],
            form_question_group=MagicMock(all=lambda: [group]),
        )

        stream, skipped = generate_xlsform(form)
        self.assertEqual(skipped, [])

        wb = openpyxl.load_workbook(stream)

        # Verify sheets
        self.assertEqual(wb.sheetnames, ["survey", "choices", "settings"])

        # Check Survey Sheet Headers
        ws_survey = wb["survey"]
        survey_rows = list(ws_survey.iter_rows(values_only=True))
        headers = survey_rows[0]
        self.assertIn("type", headers)
        self.assertIn("name", headers)
        self.assertIn("label::en", headers)
        self.assertIn("label::fr", headers)

        # Check Choices Sheet
        ws_choices = wb["choices"]
        choice_rows = list(ws_choices.iter_rows(values_only=True))
        self.assertEqual(choice_rows[0][:3], ("list_name", "name", "label"))
        self.assertEqual(choice_rows[1][:3], ("option_gender", "m", "Male"))
        self.assertEqual(choice_rows[2][:3], ("option_gender", "f", "Female"))

        # Check Settings Sheet
        ws_settings = wb["settings"]
        settings_rows = list(ws_settings.iter_rows(values_only=True))
        self.assertEqual(
            settings_rows[0],
            ("form_title", "form_id", "version", "default_language"),
        )
        self.assertEqual(
            settings_rows[1], ("Census Survey", "form_99", "3", "en")
        )

    def test_build_relevant_expression_single_and_multiple_options(self):
        qmap = {
            1: {"name": "cleaning_schedule", "type": QuestionTypes.option},
            2: {"name": "days_cleaned", "type": QuestionTypes.multiple_option},
        }

        # Single option
        q_single = DummyObject(
            dependency=[{"id": 1, "options": ["yes"]}],
            dependency_rule="AND",
        )
        self.assertEqual(
            _build_relevant_expression(q_single, qmap),
            "selected(${cleaning_schedule}, 'yes')",
        )

        # Multiple options (any-match OR)
        q_multi = DummyObject(
            dependency=[{"id": 2, "options": ["mon", "tue"]}],
            dependency_rule="AND",
        )
        self.assertEqual(
            _build_relevant_expression(q_multi, qmap),
            "(selected(${days_cleaned}, 'mon') or selected(${days_cleaned}, 'tue'))",  # noqa
        )

    def test_build_relevant_expression_min_max_equal_notequal(self):
        qmap = {
            1: {"name": "staff_count", "type": QuestionTypes.number},
            2: {"name": "status", "type": QuestionTypes.input},
        }

        q_min = DummyObject(dependency=[{"id": 1, "min": 4}])
        self.assertEqual(
            _build_relevant_expression(q_min, qmap), "${staff_count} >= 4"
        )

        q_max = DummyObject(dependency=[{"id": 1, "max": 6}])
        self.assertEqual(
            _build_relevant_expression(q_max, qmap), "${staff_count} <= 6"
        )

        q_equal = DummyObject(dependency=[{"id": 2, "equal": "active"}])
        self.assertEqual(
            _build_relevant_expression(q_equal, qmap), "${status} = 'active'"
        )

        q_notequal = DummyObject(dependency=[{"id": 2, "notEqual": "pending"}])
        self.assertEqual(
            _build_relevant_expression(q_notequal, qmap),
            "${status} != 'pending' and string-length(${status}) > 0",
        )

    def test_build_relevant_expression_combinators_and_unresolvable_id(self):
        qmap = {
            10: {"name": "cleaning_schedule", "type": QuestionTypes.option},
            20: {"name": "staff_count", "type": QuestionTypes.number},
        }

        # AND rule (default)
        q_and = DummyObject(
            dependency=[
                {"id": 10, "options": ["yes"]},
                {"id": 20, "max": 6},
            ],
            dependency_rule="AND",
        )
        self.assertEqual(
            _build_relevant_expression(q_and, qmap),
            "selected(${cleaning_schedule}, 'yes') and ${staff_count} <= 6",
        )

        # OR rule
        q_or = DummyObject(
            dependency=[
                {"id": 10, "options": ["yes"]},
                {"id": 20, "max": 6},
            ],
            dependency_rule="OR",
        )
        self.assertEqual(
            _build_relevant_expression(q_or, qmap),
            "selected(${cleaning_schedule}, 'yes') or ${staff_count} <= 6",
        )

        # Unresolvable ID skipped gracefully
        q_unresolved = DummyObject(
            dependency=[
                {"id": 10, "options": ["yes"]},
                {"id": 999, "options": ["deleted"]},
            ],
            dependency_rule="AND",
        )
        self.assertEqual(
            _build_relevant_expression(q_unresolved, qmap),
            "selected(${cleaning_schedule}, 'yes')",
        )

    def test_build_relevant_expression_edge_cases(self):
        qmap = {1: {"name": "age", "type": QuestionTypes.number}}

        # None / empty / invalid types
        self.assertEqual(
            _build_relevant_expression(DummyObject(dependency=None), qmap), ""
        )
        self.assertEqual(
            _build_relevant_expression(
                DummyObject(dependency="invalid"), qmap
            ),
            "",
        )
        self.assertEqual(
            _build_relevant_expression(DummyObject(dependency=[]), qmap), ""
        )
        self.assertEqual(
            _build_relevant_expression(
                DummyObject(dependency=["not_a_dict"]), qmap
            ),
            "",
        )

        # Options with empty list
        q_empty_opts = DummyObject(dependency=[{"id": 1, "options": []}])
        self.assertEqual(_build_relevant_expression(q_empty_opts, qmap), "")

        # Combined min and max in single dependency item
        q_range = DummyObject(dependency=[{"id": 1, "min": 18, "max": 65}])
        self.assertEqual(
            _build_relevant_expression(q_range, qmap),
            "${age} >= 18 and ${age} <= 65",
        )

    def test_build_constraint(self):
        # 1. min and max
        expr, msg = _build_constraint({"min": 1, "max": 7})
        self.assertEqual(expr, ". >= 1 and . <= 7")
        self.assertEqual(msg, "Value must be between 1 and 7")

        # 2. min only
        expr, msg = _build_constraint({"min": 5})
        self.assertEqual(expr, ". >= 5")
        self.assertEqual(msg, "Value must be at least 5")

        # 3. max only
        expr, msg = _build_constraint({"max": 10})
        self.assertEqual(expr, ". <= 10")
        self.assertEqual(msg, "Value must be at most 10")

        # 4. Zero min value edge case
        expr, msg = _build_constraint({"min": 0})
        self.assertEqual(expr, ". >= 0")
        self.assertEqual(msg, "Value must be at least 0")

        # 5. empty / None / no min or max / None values
        self.assertEqual(_build_constraint(None), (None, None))
        self.assertEqual(_build_constraint({}), (None, None))
        self.assertEqual(
            _build_constraint({"min": None, "max": None}), (None, None)
        )
        self.assertEqual(
            _build_constraint({"allowDecimal": True}), (None, None)
        )

    def test_survey_rows_t003_multilingual_addons_and_skipped_collection(
        self,
    ):
        # 1. Question with addons (should be dropped) +
        # tooltip + constraint + translations
        q_num = DummyObject(
            id=1,
            name="price",
            type=QuestionTypes.number,
            label="Price",
            required=True,
            tooltip={"text": "Enter amount in USD"},
            translations={
                "fr": {
                    "label": "Prix",
                    "tooltip": {"text": "Entrez le montant en USD"},
                },
                "es": {"label": "Precio"},
            },
            rule={"min": 1, "max": 1000},
            addon_before="$",
            addon_after="USD",
        )

        # 2. All 3 skipped types
        q_tree = DummyObject(id=2, name="tree_q", type=QuestionTypes.tree)
        q_table = DummyObject(id=3, name="table_q", type=QuestionTypes.table)
        q_auto = DummyObject(id=4, name="auto_q", type=QuestionTypes.autofield)

        group = DummyObject(
            id=10,
            name="payment_group",
            label="Payment Info",
            repeatable=False,
            translations={"fr": {"label": "Info Paiement"}},
            question_group_question=MagicMock(
                all=lambda: [q_num, q_tree, q_table, q_auto]
            ),
        )

        form = DummyObject(form_question_group=MagicMock(all=lambda: [group]))

        survey_rows, skipped = _build_survey_rows(
            form, question_map={}, lang_cols=["fr", "es"]
        )

        # Verify all 3 unsupported types collected in skipped list
        self.assertEqual(skipped, ["tree_q", "table_q", "auto_q"])

        # Verify question row
        q_row = [r for r in survey_rows if r.get("name") == "price"][0]
        self.assertEqual(q_row["type"], "integer")
        self.assertEqual(q_row["label"], "Price")
        self.assertEqual(q_row["hint"], "Enter amount in USD")
        self.assertEqual(q_row["constraint"], ". >= 1 and . <= 1000")
        self.assertEqual(
            q_row["constraint_message"], "Value must be between 1 and 1000"
        )
        self.assertEqual(q_row["label::fr"], "Prix")
        self.assertEqual(q_row["hint::fr"], "Entrez le montant en USD")
        self.assertEqual(q_row["label::es"], "Precio")

        # Verify addon_before & addon_after are NOT present in output row
        self.assertNotIn("addon_before", q_row)
        self.assertNotIn("addon_after", q_row)


def _login(client, email="admin@akvo.org", password="Test105*"):
    res = client.post(
        "/api/v1/login",
        {"email": email, "password": password},
        content_type="application/json",
    )
    return {"HTTP_AUTHORIZATION": f"Bearer {res.json().get('token')}"}


@override_settings(USE_TZ=False, TEST_ENV=True)
class FormXLSFormExportEndpointTestCase(TestCase):
    """
    Tests for GET /api/v1/manage/forms/{id}/export-xlsform (FB-014 / T-004).
    """

    def setUp(self):
        call_command("administration_seeder", "--test", 1)
        call_command("fake_organisation_seeder", "--repeat", 3)
        call_command("default_roles_seeder", "--test")
        call_command("form_seeder", "--test")
        with connection.cursor() as cur:
            for tbl in ["form", "question_group", "question", "option"]:
                cur.execute(
                    f"SELECT setval("
                    f"pg_get_serial_sequence('{tbl}', 'id'),"
                    f'(SELECT COALESCE(MAX(id), 0) FROM "{tbl}") + 1,'
                    f"false)"
                )
        self.header = _login(self.client)
        self.form = Forms.objects.first()

    def test_export_xlsform_returns_200_and_xlsx_content_type(self):
        res = self.client.get(
            f"/api/v1/manage/forms/{self.form.id}/export-xlsform",
            **self.header,
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            res["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # noqa
        )

    def test_export_xlsform_content_disposition(self):
        res = self.client.get(
            f"/api/v1/manage/forms/{self.form.id}/export-xlsform",
            **self.header,
        )
        disposition = res.get("Content-Disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn(".xlsx", disposition)

    def test_export_xlsform_unauthenticated_401(self):
        res = self.client.get(
            f"/api/v1/manage/forms/{self.form.id}/export-xlsform"
        )
        self.assertEqual(res.status_code, 401)

    def test_export_xlsform_non_existent_404(self):
        res = self.client.get(
            "/api/v1/manage/forms/999999/export-xlsform",
            **self.header,
        )
        self.assertEqual(res.status_code, 404)

    def test_export_xlsform_skipped_header_present_if_skipped(self):
        from api.v1.v1_forms.models import QuestionGroup, Questions

        g = QuestionGroup.objects.create(
            form=self.form, name="grp_skipped", order=99
        )
        Questions.objects.create(
            form=self.form,
            question_group=g,
            name="unsupported_tree",
            type=QuestionTypes.tree,
            order=1,
        )
        res = self.client.get(
            f"/api/v1/manage/forms/{self.form.id}/export-xlsform",
            **self.header,
        )
        self.assertEqual(res.status_code, 200)
        self.assertIn("X-XLSForm-Skipped", res)
        self.assertIn("unsupported_tree", res["X-XLSForm-Skipped"])
