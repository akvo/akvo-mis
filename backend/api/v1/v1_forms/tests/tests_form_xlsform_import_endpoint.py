import io
import json
import tempfile
from unittest.mock import patch

import openpyxl
from django.core.management import call_command
from django.db import connection
from django.test import TestCase
from django.test.utils import override_settings

from api.v1.v1_forms.constants import FormTypes
from api.v1.v1_forms.models import Forms
from api.v1.v1_forms.services.xlsform_export import generate_xlsform
from api.v1.v1_forms.tasks import import_xlsform_job
from api.v1.v1_jobs.constants import JobTypes, JobStatus
from api.v1.v1_jobs.models import Jobs
from api.v1.v1_users.models import SystemUser


def _login(client, email="admin@akvo.org", password="Test105*"):
    res = client.post(
        "/api/v1/login",
        {"email": email, "password": password},
        content_type="application/json",
    )
    return {"HTTP_AUTHORIZATION": f"Bearer {res.json().get('token')}"}


def _build_xlsx_file(survey_rows, settings_row=None, choices_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "survey"
    ws.append(
        [
            "type",
            "name",
            "label",
            "required",
            "hint",
            "relevant",
            "constraint",
            "appearance",
            "body::accept",
        ]
    )
    for row in survey_rows:
        ws.append(row)
    if choices_rows:
        ws_c = wb.create_sheet(title="choices")
        ws_c.append(["list_name", "name", "label"])
        for row in choices_rows:
            ws_c.append(row)
    if settings_row:
        ws_s = wb.create_sheet(title="settings")
        ws_s.append(["form_title", "form_id", "version", "default_language"])
        ws_s.append(settings_row)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    stream.name = "form.xlsx"
    return stream


@override_settings(USE_TZ=False, TEST_ENV=True)
class XLSFormImportEndpointTestCase(TestCase):
    def setUp(self):
        call_command("administration_seeder", "--test", 1)
        call_command("fake_organisation_seeder", "--repeat", 3)
        call_command("default_roles_seeder", "--test")
        call_command("form_seeder", "--test")
        with connection.cursor() as cur:
            for tbl in ["form", "question_group", "question", "option"]:
                cur.execute(
                    f"SELECT setval("
                    f"pg_get_serial_sequence('{tbl}', 'id'),"
                    f'(SELECT COALESCE(MAX(id), 0) FROM "{tbl}") + 1,'
                    f"false)"
                )
        self.header = _login(self.client)
        self.user = SystemUser.objects.filter(email="admin@akvo.org").first()
        self.existing_form = Forms.objects.first()

    def test_preflight_valid_returns_200(self):
        f = _build_xlsx_file(
            [["text", "name", "Your Name", "yes", None, None, None, None]],
            settings_row=["Household Survey", "form_1", "1", "English (en)"],
        )
        res = self.client.post(
            "/api/v1/manage/forms/import/xlsform/preflight",
            {"file": f},
            **self.header,
        )
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertTrue(data["valid"])
        self.assertEqual(data["question_count"], 1)
        self.assertEqual(data["form"]["name"], "Household Survey")

    def test_preflight_missing_file_returns_400(self):
        res = self.client.post(
            "/api/v1/manage/forms/import/xlsform/preflight",
            {},
            **self.header,
        )
        self.assertEqual(res.status_code, 400)

    def test_preflight_invalid_extension_returns_400(self):
        f = io.BytesIO(b'{"some": "json"}')
        f.name = "form.json"
        res = self.client.post(
            "/api/v1/manage/forms/import/xlsform/preflight",
            {"file": f},
            **self.header,
        )
        self.assertEqual(res.status_code, 400)
        self.assertIn(
            "Only .xlsx or .xls files are supported", res.json()["message"]
        )

    def test_preflight_empty_survey_returns_400(self):
        f = _build_xlsx_file(
            [["calculate", "calc", "Calc", None, None, None, None, None]]
        )
        res = self.client.post(
            "/api/v1/manage/forms/import/xlsform/preflight",
            {"file": f},
            **self.header,
        )
        self.assertEqual(res.status_code, 400)
        self.assertFalse(res.json()["valid"])

    @patch("api.v1.v1_forms.views.storage.upload")
    @patch(
        "api.v1.v1_forms.views.async_task", return_value="task-test-uuid-123"
    )
    def test_import_xlsform_registration_enqueues_job_200(
        self, mock_async, mock_upload
    ):
        f = _build_xlsx_file(
            [["text", "name", "Your Name", "yes", None, None, None, None]],
            settings_row=["Household Survey", "form_1", "1", "English (en)"],
        )
        res = self.client.post(
            "/api/v1/manage/forms/import/xlsform",
            {"file": f, "form_type": "registration"},
            **self.header,
        )
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertEqual(data["task_id"], "task-test-uuid-123")
        job = Jobs.objects.get(id=data["job_id"])
        self.assertEqual(job.type, JobTypes.import_form)
        self.assertEqual(job.info["form_type"], "registration")
        self.assertIsNone(job.info["parent_id"])

    def test_import_xlsform_monitoring_missing_parent_returns_400(self):
        f = _build_xlsx_file(
            [["text", "name", "Your Name", "yes", None, None, None, None]],
        )
        res = self.client.post(
            "/api/v1/manage/forms/import/xlsform",
            {"file": f, "form_type": "monitoring"},
            **self.header,
        )
        self.assertEqual(res.status_code, 400)
        self.assertIn(
            "parent_id is required for monitoring forms",
            res.json()["message"],
        )

    @patch("api.v1.v1_forms.views.storage.upload")
    @patch(
        "api.v1.v1_forms.views.async_task", return_value="task-test-uuid-456"
    )
    def test_import_xlsform_monitoring_with_parent_200(
        self, mock_async, mock_upload
    ):
        f = _build_xlsx_file(
            [["text", "name", "Your Name", "yes", None, None, None, None]],
        )
        res = self.client.post(
            "/api/v1/manage/forms/import/xlsform",
            {
                "file": f,
                "form_type": "monitoring",
                "parent_id": self.existing_form.id,
            },
            **self.header,
        )
        self.assertEqual(res.status_code, 200)
        data = res.json()
        job = Jobs.objects.get(id=data["job_id"])
        self.assertEqual(job.info["form_type"], "monitoring")
        self.assertEqual(job.info["parent_id"], self.existing_form.id)

    @patch("api.v1.v1_forms.views.storage.upload")
    @patch("api.v1.v1_forms.tasks.download")
    def test_round_trip_export_and_import_endpoint(
        self, mock_download, mock_upload
    ):
        # 1. Export seeded form to XLSForm binary
        stream, skipped = generate_xlsform(self.existing_form)
        self.assertIsNotNone(stream)

        # 2. Preflight the generated workbook
        stream.seek(0)
        stream.name = "exported_form.xlsx"
        res_preflight = self.client.post(
            "/api/v1/manage/forms/import/xlsform/preflight",
            {"file": stream},
            **self.header,
        )
        self.assertEqual(res_preflight.status_code, 200)
        self.assertTrue(res_preflight.json()["valid"])

        # 3. Post to import endpoint
        stream.seek(0)
        res_import = self.client.post(
            "/api/v1/manage/forms/import/xlsform",
            {"file": stream, "form_type": "registration"},
            **self.header,
        )
        self.assertEqual(res_import.status_code, 200)
        job_id = res_import.json()["job_id"]

        # 4. Mock download storage to return local temp file and run job
        stream.seek(0)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tf.write(stream.getvalue())
            tf.flush()
            temp_path = tf.name

        mock_download.return_value = temp_path
        import_xlsform_job(job_id)

        job = Jobs.objects.get(id=job_id)
        self.assertEqual(job.status, JobStatus.done)
        res_job = json.loads(job.result)
        new_form_id = res_job["form_id"]
        new_form = Forms.objects.get(id=new_form_id)
        self.assertEqual(new_form.name, self.existing_form.name)
        self.assertEqual(new_form.type, FormTypes.registration)
        self.assertGreater(new_form.form_questions.count(), 0)
