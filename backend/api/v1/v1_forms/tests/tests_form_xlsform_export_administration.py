import openpyxl
from django.core.management import call_command
from django.db import connection
from django.test import TestCase
from django.test.utils import override_settings

from api.v1.v1_forms.constants import QuestionTypes
from api.v1.v1_forms.models import Forms, QuestionGroup, Questions
from api.v1.v1_forms.services.xlsform_export import (
    generate_administration_csv,
    generate_xlsform,
)
from api.v1.v1_profile.models import Administration, Levels
from api.v1.v1_users.models import SystemUser


def _login(client, email="admin@akvo.org", password="Test105*"):
    res = client.post(
        "/api/v1/login",
        {"email": email, "password": password},
        content_type="application/json",
    )
    return {"HTTP_AUTHORIZATION": f"Bearer {res.json().get('token')}"}


@override_settings(USE_TZ=False, TEST_ENV=True)
class FormAdministrationCSVExportEndpointTestCase(TestCase):
    """
    Tests for GET /api/v1/manage/forms/{id}/administration-csv (FB-014).
    """

    def setUp(self):
        call_command("administration_seeder", "--test", 1)
        call_command("fake_organisation_seeder", "--repeat", 3)
        call_command("default_roles_seeder", "--test")
        call_command("form_seeder", "--test")
        with connection.cursor() as cur:
            for tbl in ["form", "question_group", "question", "option"]:
                cur.execute(
                    f"SELECT setval("
                    f"pg_get_serial_sequence('{tbl}', 'id'),"
                    f'(SELECT COALESCE(MAX(id), 0) FROM "{tbl}") + 1,'
                    f"false)"
                )
        self.header = _login(self.client)
        self.form = Forms.objects.first()
        self.user = SystemUser.objects.filter(email="admin@akvo.org").first()

    def test_export_administration_csv_endpoint(self):
        res = self.client.get(
            f"/api/v1/manage/forms/{self.form.id}/administration-csv",
            **self.header,
        )
        self.assertEqual(res.status_code, 200)
        self.assertIn("text/csv", res["Content-Type"])
        disposition = res.get("Content-Disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn("administration.csv", disposition)

        # Check CSV content header and rows
        content = res.content.decode("utf-8")
        self.assertTrue(
            content.startswith("list_name,name,label,parent_key,level")
        )
        self.assertIn("administration", content)

    def test_export_administration_csv_unauthenticated_401(self):
        res = self.client.get(
            f"/api/v1/manage/forms/{self.form.id}/administration-csv"
        )
        self.assertEqual(res.status_code, 401)

    def test_generate_administration_csv_max_level_capping(self):
        g = QuestionGroup.objects.create(
            form=self.form, name="grp_cascade", order=98
        )
        # Create cascade question with max_level = 0 (Level 0 only)
        Questions.objects.create(
            form=self.form,
            question_group=g,
            name="casc_q",
            type=QuestionTypes.cascade,
            api={"max_level": 0},
            order=1,
        )

        csv_str = generate_administration_csv(self.form, self.user)
        rows = [
            line.split(",") for line in csv_str.strip().split("\r\n") if line
        ]
        self.assertEqual(
            rows[0], ["list_name", "name", "label", "parent_key", "level"]
        )

        level_0_ids = set(
            Administration.objects.filter(level__level=0).values_list(
                "id", flat=True
            )
        )
        level_gt0_ids = set(
            Administration.objects.filter(level__level__gt=0).values_list(
                "id", flat=True
            )
        )

        exported_names = [r[1] for r in rows[1:]]
        # Check level 0 items present
        for l0_id in level_0_ids:
            adm = Administration.objects.get(id=l0_id)
            code_or_id = adm.code if adm.code else str(adm.id)
            self.assertIn(code_or_id, exported_names)

        # Check level > 0 items excluded
        for l_gt0_id in level_gt0_ids:
            adm = Administration.objects.get(id=l_gt0_id)
            code_or_id = adm.code if adm.code else str(adm.id)
            self.assertNotIn(code_or_id, exported_names)

    def test_generate_administration_csv_multiple_cascade_takes_highest_max(
        self,
    ):
        g = QuestionGroup.objects.create(
            form=self.form, name="grp_cascades", order=97
        )
        Questions.objects.create(
            form=self.form,
            question_group=g,
            name="casc_q1",
            type=QuestionTypes.cascade,
            api={"max_level": 0},
            order=1,
        )
        Questions.objects.create(
            form=self.form,
            question_group=g,
            name="casc_q2",
            type=QuestionTypes.cascade,
            api={"max_level": 1},
            order=2,
        )

        csv_str = generate_administration_csv(self.form, self.user)
        rows = [
            line.split(",") for line in csv_str.strip().split("\r\n") if line
        ]

        level_1_ids = set(
            Administration.objects.filter(level__level=1).values_list(
                "id", flat=True
            )
        )
        level_gt1_ids = set(
            Administration.objects.filter(level__level__gt=1).values_list(
                "id", flat=True
            )
        )

        exported_names = [r[1] for r in rows[1:]]
        for l1_id in level_1_ids:
            adm = Administration.objects.get(id=l1_id)
            code_or_id = adm.code if adm.code else str(adm.id)
            self.assertIn(code_or_id, exported_names)

        for l_gt1_id in level_gt1_ids:
            adm = Administration.objects.get(id=l_gt1_id)
            code_or_id = adm.code if adm.code else str(adm.id)
            self.assertNotIn(code_or_id, exported_names)

    def test_generate_administration_csv_fallback_id_and_top_level_parent_key(
        self,
    ):
        l0 = Levels.objects.filter(level=0).first()
        top_adm = Administration.objects.create(
            name="Top Level Custom",
            level=l0,
            code=None,
        )
        sub_adm = Administration.objects.create(
            name="Sub Level Custom",
            level=l0,
            parent=top_adm,
            code=None,
        )

        csv_str = generate_administration_csv(self.form, self.user)
        rows = [
            line.split(",") for line in csv_str.strip().split("\r\n") if line
        ]

        # Find row for top_adm (id as name)
        top_row = next((r for r in rows if r[1] == str(top_adm.id)), None)
        self.assertIsNotNone(top_row)
        self.assertEqual(top_row[0], "administration")
        self.assertEqual(top_row[2], "Top Level Custom")
        self.assertEqual(top_row[3], "")  # Top level has empty parent_key
        self.assertEqual(top_row[4], "0")  # Level 0

        # Find row for sub_adm
        sub_row = next((r for r in rows if r[1] == str(sub_adm.id)), None)
        self.assertIsNotNone(sub_row)
        self.assertEqual(
            sub_row[3], str(top_adm.id)
        )  # Parent key uses parent ID
        self.assertEqual(sub_row[4], "0")

    def test_dict_object_option_other_attribute_handling(self):
        dict_payload = {
            "id": 999,
            "name": "Published Form",
            "languages": ["en"],
            "default_language": "en",
            "question_group": [
                {
                    "id": 1,
                    "name": "g1",
                    "label": "Group 1",
                    "question": [
                        {
                            "id": 101,
                            "name": "choice_q",
                            "label": "Choice Question",
                            "type": "option",
                            "options": [
                                {"id": 1, "value": "a", "label": "Option A"},
                                {
                                    "id": 2,
                                    "value": "b",
                                    "label": "Option B",
                                    "other": True,
                                },
                            ],
                        }
                    ],
                }
            ],
        }
        output, skipped = generate_xlsform(dict_payload)
        self.assertIsNotNone(output)
        self.assertEqual(skipped, [])

    def test_generate_xlsform_cascade_multilevel_expansion(self):
        dict_payload = {
            "id": 888,
            "name": "Cascade Test Form",
            "languages": ["en", "id"],
            "default_language": "en",
            "question_group": [
                {
                    "id": 1,
                    "name": "grp_loc",
                    "label": "Location Group",
                    "question": [
                        {
                            "id": 10,
                            "name": "location",
                            "label": "Select Location",
                            "type": "cascade",
                            "required": True,
                            "api": {"min_level": 0, "max_level": 2},
                            "translations": [
                                {
                                    "language": "id",
                                    "label": "Pilih Lokasi",
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output, skipped = generate_xlsform(dict_payload)
        self.assertIsNotNone(output)
        self.assertEqual(skipped, [])

        wb = openpyxl.load_workbook(output)
        ws = wb["survey"]
        headers = [cell.value for cell in ws[1]]
        name_idx = headers.index("name") + 1
        type_idx = headers.index("type") + 1
        filter_idx = headers.index("choice_filter") + 1
        relevant_idx = headers.index("relevant") + 1
        label_en_idx = headers.index("label::English (en)") + 1
        label_id_idx = headers.index("label::Indonesian (id)") + 1

        # Check expanded level rows (Level 0 is omitted, begins at Level 1)
        # Row 2 is begin_group
        # Row 3 is Level 1 (Province)
        r3_type = ws.cell(row=3, column=type_idx).value
        r3_name = ws.cell(row=3, column=name_idx).value
        r3_filter = ws.cell(row=3, column=filter_idx).value
        r3_relevant = ws.cell(row=3, column=relevant_idx).value
        r3_label_en = ws.cell(row=3, column=label_en_idx).value
        r3_label_id = ws.cell(row=3, column=label_id_idx).value

        self.assertEqual(r3_type, "select_one_from_file administration.csv")
        self.assertEqual(r3_name, "location_level_1")
        self.assertEqual(r3_filter, "level = 1")
        self.assertIsNone(r3_relevant)
        self.assertIn("Province", r3_label_en)
        self.assertIn("Province", r3_label_id)

        # Row 4 is Level 2 (District)
        r4_type = ws.cell(row=4, column=type_idx).value
        r4_name = ws.cell(row=4, column=name_idx).value
        r4_filter = ws.cell(row=4, column=filter_idx).value
        r4_relevant = ws.cell(row=4, column=relevant_idx).value
        r4_label_en = ws.cell(row=4, column=label_en_idx).value

        self.assertEqual(r4_type, "select_one_from_file administration.csv")
        self.assertEqual(r4_name, "location_level_2")
        self.assertEqual(r4_filter, "parent_key = ${location_level_1}")
        self.assertEqual(r4_relevant, "${location_level_1} != ''")
        self.assertIn("District", r4_label_en)

        # Row 5 is end_group
        r5_type = ws.cell(row=5, column=type_idx).value
        self.assertEqual(r5_type, "end_group")
